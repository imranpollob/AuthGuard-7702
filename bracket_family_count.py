#!/usr/bin/env python3
"""
bracket_family_count.py

The PUSH-only skeleton (439 EOA families) is an UPPER bound on true behavioral
diversity. This script brackets the real number with three progressively stronger
family definitions, so the paper can defend whatever "family" means:

    D1  PUSH-skeleton          (what recon already did; upper bound)      -> expect ~439
    D2  4-byte-selector set    (functions the contract exposes; behavior) -> coarser
    D3  opcode 4-gram MinHash  (near-duplicate control-flow clustering)   -> tightest

For the paper you report the RANGE and pick D3-style clustering as the family
grouping for leave-family-out (it's the one a reviewer will accept as "behavioral").

Input: the EOA-targeted bytecode. Point --xlsx at
  USENIX EIP-7702 artifact/eoa_detect/get_code/contracts_with_bytecode.xlsx
and --labels at
  USENIX EIP-7702 artifact/eoa_detect/detect_result.jsonl
so we cluster ONLY the 793 confirmed-malicious rows, not all 2685 candidates.

Deps: pip install openpyxl   (stdlib for everything else)

Usage:
  python bracket_family_count.py \
     --xlsx ".../eoa_detect/get_code/contracts_with_bytecode.xlsx" \
     --labels ".../eoa_detect/detect_result.jsonl" \
     --threshold 0.85

Report back the whole "=== BRACKET RESULT ===" block.
"""
import argparse, hashlib, json, re, sys
from collections import defaultdict, Counter

def sha(s):
    if isinstance(s, str): s = s.encode()
    return hashlib.sha256(s).hexdigest()[:16]

_PUSH1, _PUSH32 = 0x60, 0x7f
def disasm(hexstr):
    h = hexstr.lower().strip()
    if h.startswith("0x"): h = h[2:]
    if len(h) % 2: h = h[:-1]
    try: b = bytes.fromhex(h)
    except ValueError: return None, None
    ops, selectors = [], set()
    i, n = 0, len(b)
    while i < n:
        op = b[i]
        if _PUSH1 <= op <= _PUSH32:
            size = op - _PUSH1 + 1
            imm = b[i+1:i+1+size]
            ops.append("PUSH")
            # PUSH4 immediates are candidate function selectors
            if size == 4 and len(imm) == 4:
                selectors.add(imm.hex())
            i += 1 + size
        else:
            ops.append(f"{op:02x}")
            i += 1
    return ops, selectors

# ---- D1: PUSH skeleton -------------------------------------------------------
def d1_skeleton(ops): return sha(" ".join(ops))

# ---- D2: selector set --------------------------------------------------------
def d2_selectorset(selectors):
    # the sorted set of 4-byte selectors = the contract's exposed interface
    return sha(",".join(sorted(selectors))) if selectors else "NO_SELECTORS"

# ---- D3: opcode 4-gram MinHash + union-find near-dup clustering ---------------
def ngram_minhash(ops, k=4, num_perm=64):
    if len(ops) < k: grams = {" ".join(ops)}
    else: grams = {" ".join(ops[i:i+k]) for i in range(len(ops)-k+1)}
    # cheap minhash: num_perm independent hashes, keep min per permutation
    sig = []
    for p in range(num_perm):
        mn = min((hash((p, g)) & 0xffffffff) for g in grams)
        sig.append(mn)
    return tuple(sig)

def minhash_sim(a, b):
    return sum(1 for x, y in zip(a, b) if x == y) / len(a)

class UF:
    def __init__(self, n): self.p = list(range(n))
    def find(self, x):
        while self.p[x] != x:
            self.p[x] = self.p[self.p[x]]; x = self.p[x]
        return x
    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb: self.p[ra] = rb

def load_rows(xlsx_path, labels_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    def col(*names):
        for nm in names:
            if nm in header: return header.index(nm)
        return None
    ci_code = col("bytecode","runtime","code","deployedbytecode")
    ci_addr = col("address","contract","ca_address","delegate","addr","delegated_address")
    ci_chain = col("chain","chain_id","chainid","network")
    if ci_code is None:
        print("FATAL: no bytecode column found; header was:", header); sys.exit(1)
    if ci_addr is None:
        print("FATAL: no address column found; header was:", header); sys.exit(1)

    # optional: restrict to confirmed-malicious via labels file.
    # detect_result.jsonl is a JSON array dumped one element per line WITHOUT the
    # enclosing brackets, so every line but the last ends in a trailing comma and
    # json.loads() fails on it as-is -- strip that before parsing. Also prefer
    # matching on the (chain, address) pair (derived from the "path" field, e.g.
    # "output/optimism/0x....hex") over address-only, because the same delegate
    # address can recur as a non-malicious candidate on other chains in the xlsx.
    keep_pairs = set()   # {(chain, addr)}
    keep_addrs = set()   # address-only fallback
    if labels_path:
        try:
            with open(labels_path, encoding="utf-8", errors="ignore") as f:
                for ln in f:
                    ln = ln.strip()
                    if not ln: continue
                    ln_json = ln[:-1] if ln.endswith(",") else ln
                    try:
                        obj = json.loads(ln_json)
                    except Exception:
                        obj = None
                    addr_here, chain_here = None, None
                    if isinstance(obj, dict):
                        for v in obj.values():
                            if isinstance(v, str) and v.startswith("0x") and len(v) == 42:
                                addr_here = v.lower()
                        path_val = obj.get("path")
                        if isinstance(path_val, str):
                            m = re.search(r'/([A-Za-z0-9_-]+)/0x[0-9a-fA-F]{40}\.hex', path_val)
                            if m: chain_here = m.group(1).lower()
                    if addr_here:
                        keep_addrs.add(addr_here)
                        if chain_here:
                            keep_pairs.add((chain_here, addr_here))
                    # bare-address regex fallback, always runs (not gated on parse success)
                    for a in re.findall(r"0x[0-9a-fA-F]{40}", ln):
                        keep_addrs.add(a.lower())
            if keep_pairs:
                print(f"[labels] {len(keep_addrs)} unique confirmed-malicious addresses "
                      f"/ {len(keep_pairs)} unique (chain,address) instances parsed from labels file")
            else:
                print(f"[labels] {len(keep_addrs)} confirmed-malicious addresses parsed from labels file "
                      f"(no chain info found in labels file; falling back to address-only match)")
        except FileNotFoundError:
            print("[labels] file not found; clustering ALL rows in the xlsx")
            keep_pairs, keep_addrs = set(), None

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = r[ci_code] if ci_code < len(r) else None
        if not code: continue
        addr = str(r[ci_addr]).lower() if (ci_addr is not None and ci_addr < len(r) and r[ci_addr]) else None
        chain = str(r[ci_chain]).lower() if (ci_chain is not None and ci_chain < len(r) and r[ci_chain]) else "?"
        if labels_path:
            if keep_pairs:
                if addr is None or (chain, addr) not in keep_pairs:
                    continue
            elif keep_addrs is not None:
                if addr is None or addr not in keep_addrs:
                    continue
        rows.append((addr, chain, str(code)))
    print(f"[load] {len(rows)} bytecode rows selected for clustering")
    return rows

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--labels", default=None, help="detect_result.jsonl to restrict to confirmed-malicious")
    ap.add_argument("--threshold", type=float, default=0.85, help="MinHash sim to merge into same D3 family")
    args = ap.parse_args()

    rows = load_rows(args.xlsx, args.labels)
    if not rows:
        print("no rows; abort"); return

    d1, d2, sigs = [], [], []
    for addr, chain, code in rows:
        ops, sels = disasm(code)
        if ops is None:
            d1.append("BAD"); d2.append("BAD"); sigs.append(None); continue
        d1.append(d1_skeleton(ops))
        d2.append(d2_selectorset(sels))
        sigs.append(ngram_minhash(ops))

    n = len(rows)
    uf = UF(n)
    # bucket by first minhash band to avoid O(n^2); then verify within bucket
    buckets = defaultdict(list)
    for i, s in enumerate(sigs):
        if s is None: continue
        band = s[:4]  # coarse LSH band
        buckets[band].append(i)
    comparisons = 0
    for idxs in buckets.values():
        for a in range(len(idxs)):
            for b in range(a+1, len(idxs)):
                i, j = idxs[a], idxs[b]
                comparisons += 1
                if minhash_sim(sigs[i], sigs[j]) >= args.threshold:
                    uf.union(i, j)
    d3_families = len({uf.find(i) for i in range(n) if sigs[i] is not None})

    print("\n=== BRACKET RESULT ===")
    print(f"instances clustered:            {n}")
    print(f"D1 PUSH-skeleton families:      {len(set(d1))}   (upper bound; recon method)")
    print(f"D2 selector-set families:       {len(set(d2))}   (exposed-interface behavioral proxy)")
    print(f"D3 opcode-4gram families @{args.threshold}: {d3_families}   (near-dup control-flow clusters; ~{comparisons} cmps)")
    print(f"clones/family under D3:         {n/max(d3_families,1):.2f}")

    # D3 family size distribution
    fam = defaultdict(int)
    for i in range(n):
        if sigs[i] is not None: fam[uf.find(i)] += 1
    sizes = sorted(fam.values(), reverse=True)
    top10 = sizes[:10]
    print(f"D3 top-10 family sizes:         {top10}  (cover {100*sum(top10)/n:.0f}% of instances)")
    singletons = sum(1 for s in sizes if s == 1)
    print(f"D3 singleton families:          {singletons} ({100*singletons/max(len(sizes),1):.0f}% of families)")
    print("\nInterpretation:")
    print("  - If D3 >> 40, leave-family-out is a real benchmark (many independent folds).")
    print("  - Report the RANGE [D3 .. D1] as 'families'; use D3 grouping for the split.")
    print("  - High singleton % = lots of one-off attacker code = good for generalization test.")
    print("=== END BRACKET ===")

if __name__ == "__main__":
    main()
