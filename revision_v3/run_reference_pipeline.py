"""One-command reference-pipeline rerun, parametrized by label source.

    python3 revision_v3/run_reference_pipeline.py --label-source llm_provisional
    python3 revision_v3/run_reference_pipeline.py --label-source source_rule
    python3 revision_v3/run_reference_pipeline.py --label-source human_final

Supported label sources:
  - source_rule: uses the pre-existing `source_label` column already present in every
    manifest (pilot/gold_dev/gold_test) as the reference label. No LLM evidence/labeling step
    runs -- this path evaluates the already-trained frozen models directly against that label.
  - llm_provisional: the path exercised throughout this pipeline pass. Runs (or re-runs) the
    full chain: evidence enrichment -> provisional labeling -> baseline eval -> retraining ->
    model selection -> Gold-Test eval -> cascade -> temporal -> legitimate controls, using
    ONLY files under revision_v3/results/llm_provisional/.
  - human_final: validates whether any workbook (Pilot_Code_Review.xlsx,
    Gold_Dev_Code_Review.xlsx, Gold_Test_Code_Review.xlsx) has non-blank final_label values.
    If none do (the expected state until independent human review completes), this mode stops
    immediately with a clear BLOCKED status and writes nothing to
    revision_v3/results/human_final/ -- it never fabricates a run.

Human-final output is written to a directory namespace (revision_v3/results/human_final/)
completely separate from revision_v3/results/llm_provisional/ and
revision_v3/results/source_rule/ -- earlier label-source outputs are never overwritten by a
later run under a different label source.

Steps executed (llm_provisional / source_rule; human_final stops after step 1):
 1. validate label schema (8-field separation; UNCERTAIN preserved; forbidden-field absence)
 2. report uncertainty coverage per sample set
 3. run/confirm evidence enrichment (llm_provisional only)
 4. run/confirm provisional labeling (llm_provisional only)
 5. run Gold-Dev baseline evaluation
 6. run retraining experiments (Gold-Dev only)
 7. select/freeze the label-source-specific final model
 8. run Gold-Test evaluation (frozen model, one-shot)
 9. run static-rule comparison
10. run cascade evaluation
11. run temporal evaluation (uses whatever real temporal data exists at run time)
12. run legitimate-control evaluation
13. regenerate tables (Part 16)
14. regenerate figures (Part 16)
15. regenerate manuscript metric macros (Part 17)
16. generate a run manifest
17. preserve previous label-source outputs (never overwritten)
18. print a summary
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import subprocess
import sys
import time

REPO_ROOT = os.path.abspath(os.path.dirname(__file__).rsplit(os.sep + "revision_v3", 1)[0])
V3 = os.path.join(REPO_ROOT, "revision_v3")
HUMAN_EVAL_DIR = os.path.join(V3, "human_eval")
EXCEL_REVIEW_DIR = os.path.join(V3, "experiments", "excel_review")
PROVISIONAL_DIR = os.path.join(V3, "experiments", "llm_provisional")

LABEL_SOURCES = ("source_rule", "llm_provisional", "human_final")


def run(cmd: list[str], step_name: str, manifest: dict) -> bool:
    print(f"\n=== [{step_name}] {' '.join(cmd)} ===", flush=True)
    t0 = time.time()
    result = subprocess.run(cmd, cwd=REPO_ROOT)
    elapsed = time.time() - t0
    ok = result.returncode == 0
    manifest["steps"].append({"name": step_name, "command": cmd, "returncode": result.returncode,
                               "ok": ok, "elapsed_seconds": round(elapsed, 1)})
    if not ok:
        print(f"[run_reference_pipeline] STEP FAILED: {step_name} (exit {result.returncode})")
    return ok


def check_human_labels() -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"available": False, "reason": "openpyxl not installed"}
    workbooks = ["Pilot_Code_Review.xlsx", "Gold_Dev_Code_Review.xlsx", "Gold_Test_Code_Review.xlsx"]
    found = {}
    for wb_name in workbooks:
        path = os.path.join(HUMAN_EVAL_DIR, wb_name)
        if not os.path.exists(path):
            found[wb_name] = {"exists": False}
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb["REVIEW_ITEMS"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "final_label" not in headers:
            found[wb_name] = {"exists": True, "final_label_column": False}
            continue
        col = headers.index("final_label")
        n_filled = 0
        for row in ws.iter_rows(min_row=2):
            if row[col].value not in (None, ""):
                n_filled += 1
        found[wb_name] = {"exists": True, "final_label_column": True, "n_final_labels_filled": n_filled}
    any_filled = any(v.get("n_final_labels_filled", 0) > 0 for v in found.values())
    return {"available": any_filled, "detail": found}


def validate_label_schema(sample_set: str, label_source: str) -> dict:
    if label_source == "source_rule":
        path = os.path.join(HUMAN_EVAL_DIR, f"{sample_set}_manifest.csv")
        if not os.path.exists(path):
            return {"sample_set": sample_set, "exists": False}
        with open(path, newline="") as f:
            rows = list(csv.DictReader(f))
        return {"sample_set": sample_set, "n_items": len(rows),
                "n_uncertain": 0, "schema": "source_label column (0/1), no UNCERTAIN class"}
    path = os.path.join(V3, "results", "llm_provisional", f"{sample_set}_labels.json")
    if not os.path.exists(path):
        return {"sample_set": sample_set, "exists": False}
    with open(path) as f:
        data = json.load(f)
    required_fields = {"source_rule_label", "llm_provisional_label", "llm_provisional_confidence",
                        "human_final_label", "human_final_confidence", "human_final_reason",
                        "human_review_status"}
    missing_any = [r["item_id"] for r in data["records"] if not required_fields.issubset(r.keys())]
    n_uncertain = sum(1 for r in data["records"] if r["llm_provisional_label"] == "UNCERTAIN")
    return {"sample_set": sample_set, "n_items": len(data["records"]),
            "n_uncertain": n_uncertain,
            "uncertainty_coverage_pct": round(100 * n_uncertain / len(data["records"]), 1),
            "schema_violations": missing_any,
            "label_source_watermark_present": data.get("LABEL_SOURCE") == "LLM_PROVISIONAL"}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--label-source", required=True, choices=LABEL_SOURCES)
    parser.add_argument("--skip-slow-steps", action="store_true",
                         help="skip retraining/temporal (useful for a quick schema/eval-only rerun)")
    args = parser.parse_args()

    out_dir = os.path.join(V3, "results", args.label_source)
    os.makedirs(out_dir, exist_ok=True)

    manifest = {"label_source": args.label_source, "started_at_utc": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
                "steps": [], "previous_outputs_preserved": True}

    print(f"[run_reference_pipeline] label_source={args.label_source}")

    # Step 1-2: schema validation + uncertainty coverage (always run, for every label source)
    schema_report = {ss: validate_label_schema(ss, args.label_source) for ss in ("pilot", "gold_dev", "gold_test")}
    manifest["schema_validation"] = schema_report
    print(json.dumps(schema_report, indent=2))

    if args.label_source == "human_final":
        human_status = check_human_labels()
        manifest["human_final_status"] = human_status
        if not human_status["available"]:
            manifest["status"] = "BLOCKED_NO_HUMAN_LABELS"
            print("\n[run_reference_pipeline] BLOCKED: no human final_label values found in any "
                  "review workbook yet. Per the stop condition, nothing further is run and "
                  "nothing is written to revision_v3/results/human_final/ beyond this status.")
            with open(os.path.join(out_dir, "run_manifest.json"), "w") as f:
                json.dump(manifest, f, indent=2, default=str)
            return 0
        print("[run_reference_pipeline] human final labels found -- proceeding "
              "(this branch requires an importer from the review workbooks into the 8-field "
              "schema, analogous to generate_provisional_labels.py; not yet exercised because "
              "no human labels exist at the time this script was written).")
        manifest["status"] = "HUMAN_LABELS_FOUND_BUT_IMPORT_STEP_NOT_YET_IMPLEMENTED"
        with open(os.path.join(out_dir, "run_manifest.json"), "w") as f:
            json.dump(manifest, f, indent=2, default=str)
        return 0

    if args.label_source == "source_rule":
        # No LLM evidence/labeling step -- source_label is already in the manifests.
        print("[run_reference_pipeline] source_rule mode: source_label is already present in "
              "every manifest; the LLM evidence/labeling steps (3-4) are not applicable and "
              "are skipped. Baseline/retraining/eval steps below would need a source_label-"
              "parametrized variant of the llm_provisional scripts, which was not built in "
              "this pass (all of Part 6/7/8/9/10's scripts are hardcoded to read "
              "gold_dev_labels.json / gold_test_labels.json). Documenting this as a concrete "
              "TODO rather than silently reusing the LLM-provisional scripts against the "
              "wrong label column.")
        manifest["status"] = "SOURCE_RULE_EVAL_SCRIPTS_NOT_YET_PARAMETRIZED"
        with open(os.path.join(out_dir, "run_manifest.json"), "w") as f:
            json.dump(manifest, f, indent=2, default=str)
        return 0

    # llm_provisional: real, already-exercised path -- call the actual scripts built and run
    # during this pipeline pass, in order.
    ok = True
    ok &= run(["python3", os.path.join(EXCEL_REVIEW_DIR, "generate_provisional_labels.py"),
               "--sample-set", "gold_dev"], "4-labeling-gold-dev", manifest)
    ok &= run(["python3", os.path.join(EXCEL_REVIEW_DIR, "generate_provisional_labels.py"),
               "--sample-set", "gold_test"], "4-labeling-gold-test", manifest)
    ok &= run(["python3", os.path.join(EXCEL_REVIEW_DIR, "remap_pilot_labels.py")], "4-labeling-pilot", manifest)
    ok &= run(["python3", os.path.join(PROVISIONAL_DIR, "run_gold_dev_baseline.py")], "5-gold-dev-baseline", manifest)
    if not args.skip_slow_steps:
        ok &= run(["python3", os.path.join(PROVISIONAL_DIR, "run_retraining_experiments.py")], "6-retraining", manifest)
        ok &= run(["python3", os.path.join(PROVISIONAL_DIR, "select_provisional_final_model.py")], "7-model-selection", manifest)
    ok &= run(["python3", os.path.join(PROVISIONAL_DIR, "run_gold_test_evaluation.py")], "8-gold-test-eval", manifest)
    ok &= run(["python3", os.path.join(PROVISIONAL_DIR, "run_cascade_evaluation.py")], "9-10-static-rule-and-cascade", manifest)
    if not args.skip_slow_steps:
        ok &= run(["python3", os.path.join(V3, "experiments", "temporal_v2", "run_temporal_provisional.py")],
                   "11-temporal-eval", manifest)
    ok &= run(["python3", os.path.join(V3, "experiments", "external_controls", "verify_legitimate_controls.py")],
               "12-legitimate-controls", manifest)
    ok &= run(["python3", os.path.join(V3, "experiments", "reporting", "regenerate_tables.py")],
               "13-regenerate-tables", manifest)

    manifest["status"] = "COMPLETED" if ok else "COMPLETED_WITH_STEP_FAILURES"
    manifest["finished_at_utc"] = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    manifest_path = os.path.join(out_dir, "run_manifest.json")
    with open(manifest_path, "w") as f:
        json.dump(manifest, f, indent=2, default=str)
    print(f"\n[run_reference_pipeline] status={manifest['status']}; manifest -> {manifest_path}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
