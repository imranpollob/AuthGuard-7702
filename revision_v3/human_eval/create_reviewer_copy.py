#!/usr/bin/env python3
"""Creates a clean, contributor-specific copy of a review workbook (Pilot / Gold-Dev /
Gold-Test). Preserves all evidence and the LLM preliminary review; fills in the contributor's
name automatically; leaves contributor decision columns blank; keeps lead-author final-decision
columns protected (read-only) so a contributor can't accidentally edit them. Requires no
macros -- contributors just open and edit the file in any normal spreadsheet application.

Usage:
    python3 revision_v3/human_eval/create_reviewer_copy.py \\
      --input revision_v3/human_eval/Pilot_Review.xlsx \\
      --reviewer "Contributor Name" \\
      --output revision_v3/human_eval/reviewer_copies/Pilot_Review_Contributor_Name.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "experiments", "excel_review"))
from excel_builder import ALL_ITEM_COLUMNS, LEAD_AUTHOR_COLUMNS, protect_lead_author_columns  # noqa: E402

ITEMS_SHEET_CANDIDATES = ["PILOT_ITEMS", "GOLD_DEV_ITEMS", "GOLD_TEST_ITEMS"]


def _find_items_sheet(wb) -> str:
    for name in ITEMS_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return name
    raise ValueError(f"no known items sheet found; expected one of {ITEMS_SHEET_CANDIDATES}")


def create_reviewer_copy(input_path: str, reviewer_name: str, output_path: str) -> None:
    wb = openpyxl.load_workbook(input_path)
    items_sheet_name = _find_items_sheet(wb)
    ws = wb[items_sheet_name]

    contributor_name_col = ALL_ITEM_COLUMNS.index("contributor_name") + 1
    n_rows = ws.max_row - 1  # minus header

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=contributor_name_col, value=reviewer_name)

    # re-apply protection (loading/saving can drop it) and double check lead-author columns
    # are locked and everything else stays editable.
    protect_lead_author_columns(ws, n_rows=n_rows)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"wrote {output_path} for reviewer '{reviewer_name}' ({n_rows} items)")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--reviewer", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    create_reviewer_copy(args.input, args.reviewer, args.output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
