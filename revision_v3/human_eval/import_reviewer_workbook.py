#!/usr/bin/env python3
"""Imports one completed contributor Excel workbook into the master adjudication workbook.

Validates:
  - item IDs in the contributor file match the master's item set exactly
  - reviewer name is present and non-empty
  - every filled-in label is one of SAFE/UNSAFE/UNCERTAIN
  - every filled-in reason category is a member of the taxonomy AND consistent with the
    label it was recorded against (e.g. an UNSAFE reason paired with a SAFE label is rejected)
  - missing decisions are reported (not silently ignored)
  - the SAME reviewer name has not already been imported (duplicate-reviewer detection)

Does NOT compute a majority-vote final label -- it only appends this contributor's raw
responses as a new section and refreshes the disagreement_summary/discussion_notes columns.
The lead author reads everything and fills in final_label manually.

Usage:
    python3 revision_v3/human_eval/import_reviewer_workbook.py \\
      --master revision_v3/human_eval/Pilot_Master_Adjudication.xlsx \\
      --contributor-file revision_v3/human_eval/reviewer_copies/Pilot_Review_Contributor_Name.xlsx
"""
from __future__ import annotations

import argparse
import datetime
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "experiments", "excel_review"))
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from excel_builder import ALL_ITEM_COLUMNS, REASONS_BY_LABEL  # noqa: E402
from taxonomy import PRIMARY_LABELS  # noqa: E402

CONTRIBUTOR_SUFFIX_COLUMNS = ["label", "reason_category", "confidence", "rationale"]
ITEMS_SHEET_CANDIDATES = ["PILOT_ITEMS", "GOLD_DEV_ITEMS", "GOLD_TEST_ITEMS"]
MASTER_ITEMS_SHEET = "MASTER_ITEMS"


class ImportValidationError(Exception):
    pass


def _find_source_items_sheet(wb):
    for name in ITEMS_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    raise ImportValidationError(f"contributor file has no recognized items sheet ({ITEMS_SHEET_CANDIDATES})")


def _read_contributor_rows(ws) -> tuple[str, dict]:
    headers = [c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}
    rows = {}
    reviewer_name = None
    for r in range(2, ws.max_row + 1):
        item_id = ws.cell(row=r, column=col_idx["item_id"] + 1).value
        if not item_id:
            continue
        name = ws.cell(row=r, column=col_idx["contributor_name"] + 1).value
        if name:
            reviewer_name = name
        rows[item_id] = {
            "label": ws.cell(row=r, column=col_idx["contributor_label"] + 1).value,
            "reason_category": ws.cell(row=r, column=col_idx["contributor_reason_category"] + 1).value,
            "confidence": ws.cell(row=r, column=col_idx["contributor_confidence"] + 1).value,
            "rationale": ws.cell(row=r, column=col_idx["contributor_rationale"] + 1).value,
            "agree_with_llm": ws.cell(row=r, column=col_idx["agree_with_llm"] + 1).value,
            "important_evidence": ws.cell(row=r, column=col_idx["important_evidence"] + 1).value,
            "discussion_notes": ws.cell(row=r, column=col_idx["questions_or_discussion_notes"] + 1).value,
        }
    if reviewer_name is None:
        raise ImportValidationError("contributor_name column is empty for every row -- cannot identify the reviewer")
    return reviewer_name, rows


def validate_contributor_data(reviewer_name: str, rows: dict, master_item_ids: set) -> list[str]:
    warnings = []
    contributor_item_ids = set(rows.keys())
    if contributor_item_ids != master_item_ids:
        missing = master_item_ids - contributor_item_ids
        extra = contributor_item_ids - master_item_ids
        raise ImportValidationError(
            f"item ID mismatch between contributor file and master workbook. "
            f"Missing from contributor file: {sorted(missing)[:5]}{'...' if len(missing) > 5 else ''}. "
            f"Not recognized (not in master): {sorted(extra)[:5]}{'...' if len(extra) > 5 else ''}."
        )
    if not reviewer_name or not str(reviewer_name).strip():
        raise ImportValidationError("reviewer name is blank")

    for item_id, data in rows.items():
        label = data["label"]
        reason = data["reason_category"]
        if label is None or str(label).strip() == "":
            warnings.append(f"{item_id}: MISSING decision (no contributor_label filled in)")
            continue
        if label not in PRIMARY_LABELS:
            raise ImportValidationError(f"{item_id}: invalid label '{label}' (must be one of {PRIMARY_LABELS})")
        if reason:
            allowed = REASONS_BY_LABEL.get(label, [])
            if reason not in allowed:
                raise ImportValidationError(
                    f"{item_id}: reason_category '{reason}' is not valid for label '{label}' "
                    f"(allowed: {allowed})"
                )
    return warnings


def check_duplicate_reviewer(wb, reviewer_name: str) -> None:
    if "_IMPORT_LOG" not in wb.sheetnames:
        return
    log = wb["_IMPORT_LOG"]
    for row in range(2, log.max_row + 1):
        existing = log.cell(row=row, column=1).value
        if existing and str(existing).strip().lower() == str(reviewer_name).strip().lower():
            raise ImportValidationError(
                f"reviewer '{reviewer_name}' has already been imported into this master workbook "
                "(duplicate-reviewer import is not allowed -- if this is a resubmission, remove "
                "the prior section manually first)"
            )


def import_contributor(master_path: str, contributor_path: str) -> dict:
    contributor_wb = openpyxl.load_workbook(contributor_path)
    source_ws = _find_source_items_sheet(contributor_wb)
    reviewer_name, contributor_rows = _read_contributor_rows(source_ws)

    master_wb = openpyxl.load_workbook(master_path)
    master_ws = master_wb[MASTER_ITEMS_SHEET]
    master_headers = [c.value for c in master_ws[1]]
    master_item_ids = {master_ws.cell(row=r, column=1).value for r in range(2, master_ws.max_row + 1)}

    check_duplicate_reviewer(master_wb, reviewer_name)
    warnings = validate_contributor_data(reviewer_name, contributor_rows, master_item_ids)

    # insert 4 new columns for this reviewer, right before disagreement_summary
    insert_before = master_headers.index("disagreement_summary") + 1
    safe_name = "".join(c if c.isalnum() else "_" for c in reviewer_name)
    new_headers = [f"{safe_name}_{suffix}" for suffix in CONTRIBUTOR_SUFFIX_COLUMNS]

    master_ws.insert_cols(insert_before, amount=4)
    for i, header in enumerate(new_headers):
        cell = master_ws.cell(row=1, column=insert_before + i, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        master_ws.column_dimensions[get_column_letter(insert_before + i)].width = 30

    item_id_to_row = {master_ws.cell(row=r, column=1).value: r for r in range(2, master_ws.max_row + 1)}
    discussion_col = master_headers.index("discussion_notes_combined") + 1
    if insert_before <= discussion_col:
        discussion_col += 4  # shifted by the inserted columns

    for item_id, data in contributor_rows.items():
        r = item_id_to_row[item_id]
        for i, suffix in enumerate(CONTRIBUTOR_SUFFIX_COLUMNS):
            cell = master_ws.cell(row=r, column=insert_before + i, value=data.get(suffix))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if data.get("discussion_notes"):
            existing = master_ws.cell(row=r, column=discussion_col).value or ""
            note = f"[{reviewer_name}] {data['discussion_notes']}"
            master_ws.cell(row=r, column=discussion_col,
                           value=(existing + " | " + note) if existing else note)

    # recompute disagreement_summary across all *_label CONTRIBUTOR columns -- explicitly
    # excluding llm_proposed_label (advisory-only, not a human reviewer) and final_label (the
    # lead author's decision, not a contributor input) even though both also end in "_label".
    refreshed_headers = [c.value for c in master_ws[1]]
    NON_CONTRIBUTOR_LABEL_COLUMNS = {"llm_proposed_label", "final_label"}
    label_cols = [
        i + 1 for i, h in enumerate(refreshed_headers)
        if h and h.endswith("_label") and h not in NON_CONTRIBUTOR_LABEL_COLUMNS
    ]
    disagreement_col = refreshed_headers.index("disagreement_summary") + 1
    for r in range(2, master_ws.max_row + 1):
        labels = [master_ws.cell(row=r, column=c).value for c in label_cols]
        labels = [l for l in labels if l]
        if not labels:
            summary = "(no contributor responses imported yet)"
        elif len(set(labels)) == 1:
            summary = f"unanimous: {labels[0]} ({len(labels)} reviewer(s))"
        else:
            from collections import Counter
            counts = Counter(labels)
            summary = "DISAGREEMENT: " + ", ".join(f"{k}x{v}" for k, v in counts.items())
        master_ws.cell(row=r, column=disagreement_col, value=summary)

    if "_IMPORT_LOG" not in master_wb.sheetnames:
        master_wb.create_sheet("_IMPORT_LOG").append(
            ["reviewer_name", "source_file", "imported_at_utc", "n_items_imported"])
    log = master_wb["_IMPORT_LOG"]
    log.append([reviewer_name, os.path.basename(contributor_path),
               datetime.datetime.now(datetime.timezone.utc).isoformat(), len(contributor_rows)])

    master_wb.save(master_path)

    return {
        "reviewer_name": reviewer_name,
        "n_items_imported": len(contributor_rows),
        "warnings": warnings,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--master", required=True)
    parser.add_argument("--contributor-file", required=True)
    args = parser.parse_args()
    try:
        result = import_contributor(args.master, args.contributor_file)
    except ImportValidationError as e:
        print(f"IMPORT REJECTED: {e}", file=sys.stderr)
        return 1
    print(f"imported {result['n_items_imported']} responses from '{result['reviewer_name']}'")
    for w in result["warnings"]:
        print(f"  WARNING: {w}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
