"""Phase 3A: Excel-based collaborative review workflow tests."""
import csv
import hashlib
import os
import sys

import openpyxl
import pytest

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
HUMAN_EVAL_DIR = os.path.join(REPO_ROOT, "revision_v3", "human_eval")
sys.path.insert(0, os.path.join(REPO_ROOT, "revision_v3", "experiments", "excel_review"))
sys.path.insert(0, HUMAN_EVAL_DIR)

# Recorded immediately before any Phase 3A work began (see PHASE3A_EXCEL_REVIEW_PREPARATION.md).
BASELINE_MANIFEST_MD5 = {
    "pilot_manifest.csv": "2d5d84963daaa1f9e4fdb852f8c4e7cc",
    "gold_dev_manifest.csv": "a8fac81f6e2f320c23fe796d42ca6783",
    "gold_test_manifest.csv": "543f3a5a69cef800eb6a6e830f729669",
    "gold_test_hashes.json": "0dbd38312d9bca1e125416f979acabb6",
}


def _md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        h.update(f.read())
    return h.hexdigest()


def _skip_if_missing(path):
    if not os.path.exists(path):
        pytest.skip(f"{path} not present in this environment")


# ---------------------------------------------------------------------------
# No resampling / manifests unchanged
# ---------------------------------------------------------------------------

def test_pilot_manifest_unchanged_since_phase2():
    path = os.path.join(HUMAN_EVAL_DIR, "pilot_manifest.csv")
    assert _md5(path) == BASELINE_MANIFEST_MD5["pilot_manifest.csv"]


def test_gold_dev_manifest_unchanged():
    path = os.path.join(HUMAN_EVAL_DIR, "gold_dev_manifest.csv")
    assert _md5(path) == BASELINE_MANIFEST_MD5["gold_dev_manifest.csv"]


def test_gold_test_manifest_unchanged():
    path = os.path.join(HUMAN_EVAL_DIR, "gold_test_manifest.csv")
    assert _md5(path) == BASELINE_MANIFEST_MD5["gold_test_manifest.csv"]


def test_gold_test_hashes_unchanged():
    path = os.path.join(HUMAN_EVAL_DIR, "gold_test_hashes.json")
    assert _md5(path) == BASELINE_MANIFEST_MD5["gold_test_hashes.json"]


def test_no_gold_review_files_created():
    """Do not begin Gold-Dev or Gold-Test review -- these files must not exist yet."""
    for fname in ("Gold_Dev_Review.xlsx", "Gold_Test_Review.xlsx",
                 "Gold_Dev_Master_Adjudication.xlsx", "Gold_Test_Master_Adjudication.xlsx"):
        assert not os.path.exists(os.path.join(HUMAN_EVAL_DIR, fname)), (
            f"{fname} exists -- Gold-Dev/Gold-Test review must not have begun"
        )


# ---------------------------------------------------------------------------
# Pilot set integrity
# ---------------------------------------------------------------------------

def _load_pilot_manifest_ids():
    path = os.path.join(HUMAN_EVAL_DIR, "pilot_manifest.csv")
    with open(path, newline="") as f:
        return [row["item_id"] for row in csv.DictReader(f)]


def test_pilot_set_has_exactly_20_items():
    assert len(_load_pilot_manifest_ids()) == 20


def test_pilot_workbook_item_ids_match_manifest():
    path = os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx")
    _skip_if_missing(path)
    manifest_ids = set(_load_pilot_manifest_ids())
    wb = openpyxl.load_workbook(path)
    ws = wb["PILOT_ITEMS"]
    workbook_ids = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert workbook_ids == manifest_ids


# ---------------------------------------------------------------------------
# Model scores / source labels absent from the workbook
# ---------------------------------------------------------------------------

def test_workbook_contains_no_forbidden_fields():
    from taxonomy import FORBIDDEN_FIELDS
    path = os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx")
    _skip_if_missing(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["PILOT_ITEMS"]
    headers = {str(c.value).lower() for c in ws[1] if c.value}
    assert not (FORBIDDEN_FIELDS & headers)


def test_workbook_cell_values_contain_no_forbidden_substrings():
    path = os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx")
    _skip_if_missing(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["PILOT_ITEMS"]
    forbidden_substrings = ["authguard_score", "calibrated_score", "raw_score", "source_label",
                            "authguard_prediction"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                lowered = cell.value.lower()
                for bad in forbidden_substrings:
                    assert bad not in lowered, f"forbidden substring '{bad}' found in cell {cell.coordinate}"


# ---------------------------------------------------------------------------
# Excel dropdown validity
# ---------------------------------------------------------------------------

def test_dropdown_values_match_taxonomy():
    from taxonomy import AGREEMENT_VALUES, ALL_REASON_CATEGORIES, CONFIDENCE_LEVELS, PRIMARY_LABELS
    path = os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx")
    _skip_if_missing(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["PILOT_ITEMS"]
    dvs = {str(dv.sqref): dv.formula1 for dv in ws.data_validations.dataValidation}
    all_formulas = " ".join(dvs.values())
    for label in PRIMARY_LABELS:
        assert label in all_formulas
    for reason in ALL_REASON_CATEGORIES:
        assert reason in all_formulas
    for level in CONFIDENCE_LEVELS:
        assert level in all_formulas
    for val in AGREEMENT_VALUES:
        assert val in all_formulas


def test_reason_categories_partition_by_label():
    from taxonomy import ALL_REASON_CATEGORIES, REASONS_BY_LABEL, SAFE_REASONS, UNCERTAIN_REASONS, UNSAFE_REASONS
    assert set(SAFE_REASONS) | set(UNSAFE_REASONS) | set(UNCERTAIN_REASONS) == set(ALL_REASON_CATEGORIES)
    assert set(SAFE_REASONS).isdisjoint(UNSAFE_REASONS)
    assert set(SAFE_REASONS).isdisjoint(UNCERTAIN_REASONS)
    assert set(UNSAFE_REASONS).isdisjoint(UNCERTAIN_REASONS)
    assert REASONS_BY_LABEL["SAFE"] == SAFE_REASONS


# ---------------------------------------------------------------------------
# Reviewer-copy generation
# ---------------------------------------------------------------------------

def test_reviewer_copy_generation_and_contributor_name_insertion(tmp_path):
    from create_reviewer_copy import create_reviewer_copy
    input_path = os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx")
    _skip_if_missing(input_path)
    out_path = str(tmp_path / "copy.xlsx")
    create_reviewer_copy(input_path, "Unit Test Reviewer", out_path)
    assert os.path.exists(out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["PILOT_ITEMS"]
    headers = [c.value for c in ws[1]]
    idx = headers.index("contributor_name") + 1
    for r in range(2, ws.max_row + 1):
        assert ws.cell(row=r, column=idx).value == "Unit Test Reviewer"


def test_final_label_columns_remain_locked_in_reviewer_copy(tmp_path):
    from create_reviewer_copy import create_reviewer_copy
    input_path = os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx")
    _skip_if_missing(input_path)
    out_path = str(tmp_path / "copy.xlsx")
    create_reviewer_copy(input_path, "Unit Test Reviewer", out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["PILOT_ITEMS"]
    headers = [c.value for c in ws[1]]
    assert ws.protection.sheet is True
    for col_name in ("final_label", "final_reason_category", "final_rationale",
                     "final_decision_date", "included_in_binary_evaluation"):
        idx = headers.index(col_name) + 1
        assert ws.cell(row=2, column=idx).protection.locked is True
    # contributor columns must remain editable
    for col_name in ("contributor_label", "contributor_rationale"):
        idx = headers.index(col_name) + 1
        assert ws.cell(row=2, column=idx).protection.locked is False


# ---------------------------------------------------------------------------
# Workbook import validation
# ---------------------------------------------------------------------------

def _make_filled_copy(tmp_path, reviewer_name, label="UNCERTAIN", reason="INSUFFICIENT_EVIDENCE"):
    from create_reviewer_copy import create_reviewer_copy
    input_path = os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx")
    out_path = str(tmp_path / f"{reviewer_name.replace(' ', '_')}.xlsx")
    create_reviewer_copy(input_path, reviewer_name, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["PILOT_ITEMS"]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=idx["contributor_label"], value=label)
        ws.cell(row=r, column=idx["contributor_reason_category"], value=reason)
        ws.cell(row=r, column=idx["contributor_confidence"], value="medium")
    wb.save(out_path)
    return out_path


@pytest.fixture
def fresh_master(tmp_path):
    from build_master_workbook import build_master_workbook
    out_path = str(tmp_path / "master.xlsx")
    build_master_workbook(
        manifest_path=os.path.join(HUMAN_EVAL_DIR, "pilot_manifest.csv"),
        evidence_dump_path=os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_evidence_dump.json"),
        llm_reviews_path=os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_llm_reviews.json"),
        output_path=out_path,
        review_set_name="Pilot",
    )
    return out_path


def test_import_valid_contributor_succeeds(tmp_path, fresh_master):
    _skip_if_missing(os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx"))
    from import_reviewer_workbook import import_contributor
    copy_path = _make_filled_copy(tmp_path, "Valid Reviewer")
    result = import_contributor(fresh_master, copy_path)
    assert result["n_items_imported"] == 20
    assert result["warnings"] == []


def test_import_rejects_invalid_label(tmp_path, fresh_master):
    _skip_if_missing(os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx"))
    from import_reviewer_workbook import ImportValidationError, import_contributor
    copy_path = _make_filled_copy(tmp_path, "Bad Label Reviewer", label="MAYBE")
    with pytest.raises(ImportValidationError):
        import_contributor(fresh_master, copy_path)


def test_import_rejects_mismatched_reason_category(tmp_path, fresh_master):
    _skip_if_missing(os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx"))
    from import_reviewer_workbook import ImportValidationError, import_contributor
    copy_path = _make_filled_copy(tmp_path, "Mismatch Reviewer", label="SAFE",
                                  reason="ARBITRARY_EXTERNAL_CALL")
    with pytest.raises(ImportValidationError):
        import_contributor(fresh_master, copy_path)


def test_import_detects_duplicate_reviewer(tmp_path, fresh_master):
    _skip_if_missing(os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx"))
    from import_reviewer_workbook import ImportValidationError, import_contributor
    copy_path = _make_filled_copy(tmp_path, "Repeat Reviewer")
    import_contributor(fresh_master, copy_path)  # first import succeeds
    with pytest.raises(ImportValidationError):
        import_contributor(fresh_master, copy_path)  # second import of same reviewer rejected


def test_import_reports_missing_decisions_as_warning(tmp_path, fresh_master):
    _skip_if_missing(os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx"))
    from create_reviewer_copy import create_reviewer_copy
    from import_reviewer_workbook import import_contributor
    input_path = os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx")
    out_path = str(tmp_path / "Partial_Reviewer.xlsx")
    create_reviewer_copy(input_path, "Partial Reviewer", out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["PILOT_ITEMS"]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, ws.max_row):  # leave the LAST row blank on purpose
        ws.cell(row=r, column=idx["contributor_label"], value="SAFE")
        ws.cell(row=r, column=idx["contributor_reason_category"], value="NO_CONCRETE_DANGEROUS_PATH_FOUND")
    wb.save(out_path)

    result = import_contributor(fresh_master, out_path)
    assert len(result["warnings"]) == 1
    assert "MISSING decision" in result["warnings"][0]


def test_no_automatic_majority_vote_final_label(tmp_path, fresh_master):
    """Importing multiple disagreeing contributors must never write anything into final_label."""
    _skip_if_missing(os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx"))
    from import_reviewer_workbook import import_contributor
    copy_a = _make_filled_copy(tmp_path, "Reviewer A", label="SAFE", reason="NO_CONCRETE_DANGEROUS_PATH_FOUND")
    copy_b = _make_filled_copy(tmp_path, "Reviewer B", label="SAFE", reason="NO_CONCRETE_DANGEROUS_PATH_FOUND")
    copy_c = _make_filled_copy(tmp_path, "Reviewer C", label="UNSAFE", reason="ARBITRARY_EXTERNAL_CALL")
    import_contributor(fresh_master, copy_a)
    import_contributor(fresh_master, copy_b)
    import_contributor(fresh_master, copy_c)

    wb = openpyxl.load_workbook(fresh_master)
    ws = wb["MASTER_ITEMS"]
    headers = [c.value for c in ws[1]]
    final_idx = headers.index("final_label") + 1
    disagreement_idx = headers.index("disagreement_summary") + 1
    for r in range(2, ws.max_row + 1):
        assert ws.cell(row=r, column=final_idx).value in (None, ""), (
            "final_label must never be auto-populated, even under a 2-1 majority"
        )
        assert ws.cell(row=r, column=disagreement_idx).value.startswith("DISAGREEMENT")


def test_disagreement_calc_excludes_llm_and_final_columns(tmp_path, fresh_master):
    """Regression test for the bug found during Phase 3A development: llm_proposed_label and
    final_label both end in '_label' and must never be counted as contributor votes."""
    _skip_if_missing(os.path.join(HUMAN_EVAL_DIR, "Pilot_Review.xlsx"))
    from import_reviewer_workbook import import_contributor
    copy_a = _make_filled_copy(tmp_path, "Solo Reviewer", label="SAFE", reason="NO_CONCRETE_DANGEROUS_PATH_FOUND")
    import_contributor(fresh_master, copy_a)

    wb = openpyxl.load_workbook(fresh_master)
    ws = wb["MASTER_ITEMS"]
    headers = [c.value for c in ws[1]]
    disagreement_idx = headers.index("disagreement_summary") + 1
    for r in range(2, ws.max_row + 1):
        value = ws.cell(row=r, column=disagreement_idx).value
        assert value.startswith("unanimous"), f"expected unanimous with 1 real reviewer, got: {value}"
        assert "1 reviewer" in value
