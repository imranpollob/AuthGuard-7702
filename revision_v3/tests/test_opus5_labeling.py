"""Tests for the Opus 5 provisional labeling pass.

Covers the invariants the labeling instruction is explicit about: label-source separation,
no LLM-to-human copying, no Gold-Test use during development, uncertainty preservation,
watermarking, output-directory separation, blinding from AuthGuard predictions (but NOT from
the static analyzer, which is deliberately visible this pass), schema completeness, manifest
immutability, and the honesty of the analysis's own coverage accounting.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import subprocess
import sys

import pytest

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
V3 = os.path.join(ROOT, "revision_v3")
OPUS5 = os.path.join(V3, "results", "llm_provisional_opus5")
PREV = os.path.join(V3, "results", "llm_provisional")
HUMAN_EVAL = os.path.join(V3, "human_eval")
SETS = ("pilot", "gold_dev", "gold_test")
LABELS = {"SAFE", "UNSAFE", "UNCERTAIN"}

sys.path.insert(0, os.path.join(V3, "experiments", "opus5_labeling"))

csv.field_size_limit(10 ** 9)

REQUIRED_FIELDS = {
    "item_id", "sample_set", "chain", "address", "source_rule_label", "source_rule_name",
    "source_rule_assessment", "previous_llm_provisional_label", "opus5_provisional_label",
    "opus5_confidence", "reason_category", "contract_purpose", "actual_implementation",
    "sensitive_entry_points", "sensitive_operations", "caller_authorization_analysis",
    "initialization_analysis", "proxy_and_upgrade_analysis", "asset_operation_analysis",
    "eip7702_specific_analysis", "static_analyzer_evidence_summary",
    "static_analyzer_verdict_assessment", "concrete_safe_controls", "concrete_unsafe_paths",
    "conflicting_evidence", "unresolved_questions", "final_rationale", "evidence_references",
    "human_final_label", "human_final_confidence", "human_final_reason", "human_review_status",
}

# Baseline hashes recorded in Phase 3A; re-checked here so this pass cannot have altered the
# frozen sampling manifests.
MANIFEST_MD5 = {
    "pilot_manifest.csv": "2d5d84963daaa1f9e4fdb852f8c4e7cc",
    "gold_dev_manifest.csv": "a8fac81f6e2f320c23fe796d42ca6783",
    "gold_test_manifest.csv": "543f3a5a69cef800eb6a6e830f729669",
    "gold_test_hashes.json": "0dbd38312d9bca1e125416f979acabb6",
}


def load(ss):
    with open(os.path.join(OPUS5, f"{ss}_reviews_opus5.json")) as f:
        return json.load(f)


@pytest.fixture(scope="module")
def all_records():
    return {ss: load(ss) for ss in SETS}


# ---------------------------------------------------------------- schema & separation ----

@pytest.mark.parametrize("ss", SETS)
def test_every_record_has_the_full_output_schema(all_records, ss):
    for r in all_records[ss]["records"]:
        missing = REQUIRED_FIELDS - set(r)
        assert not missing, f"{r['item_id']} missing {sorted(missing)}"


@pytest.mark.parametrize("ss", SETS)
def test_human_final_fields_are_never_populated(all_records, ss):
    for r in all_records[ss]["records"]:
        assert r["human_final_label"] == ""
        assert r["human_final_confidence"] == ""
        assert r["human_final_reason"] == ""
        assert r["human_review_status"] == "NOT_REVIEWED"


@pytest.mark.parametrize("ss", SETS)
def test_no_opus5_label_is_copied_into_a_human_field(all_records, ss):
    for r in all_records[ss]["records"]:
        assert r["human_final_label"] != r["opus5_provisional_label"] or r["human_final_label"] == ""
        assert r["human_final_label"] not in LABELS


@pytest.mark.parametrize("ss", SETS)
def test_label_and_confidence_values_are_valid(all_records, ss):
    for r in all_records[ss]["records"]:
        assert r["opus5_provisional_label"] in LABELS
        assert r["opus5_confidence"] in {"HIGH", "MEDIUM", "LOW"}


@pytest.mark.parametrize("ss", SETS)
def test_the_four_label_sources_stay_distinct_fields(all_records, ss):
    """source_rule / previous LLM / Opus 5 / human must each have their own field."""
    for r in all_records[ss]["records"]:
        assert r["source_rule_label"] in {"positive", "unflagged"}
        assert r["previous_llm_provisional_label"] in (LABELS | {""})
        assert "opus5_provisional_label" in r and "human_final_label" in r


# ---------------------------------------------------------------- watermarking -----------

@pytest.mark.parametrize("ss", SETS)
def test_json_and_csv_carry_the_required_banner(all_records, ss):
    d = all_records[ss]
    assert d["LABEL_SOURCE"] == "LLM_PROVISIONAL_OPUS5"
    assert d["STATIC_ANALYZER_EVIDENCE"] == "VISIBLE"
    assert d["STATUS"] == "PROVISIONAL_PENDING_HUMAN_REVIEW"
    head = open(os.path.join(OPUS5, f"{ss}_labels_opus5.csv")).read(400)
    assert "LABEL_SOURCE=LLM_PROVISIONAL_OPUS5" in head
    assert "STATIC_ANALYZER_EVIDENCE=VISIBLE" in head
    assert "STATUS=PROVISIONAL_PENDING_HUMAN_REVIEW" in head


def test_reports_carry_the_provisional_watermark():
    for name in ("OPUS5_LABEL_COMPARISON_REPORT.md", "OPUS5_LABEL_QUALITY_REPORT.md",
                 "OPUS5_LABELING_CONTEXT_SUMMARY.md"):
        text = open(os.path.join(V3, "reports", name)).read()
        assert "LLM_PROVISIONAL_OPUS5" in text
        assert "PROVISIONAL" in text


def test_quality_report_states_the_static_analyzer_comparison_limitation():
    text = open(os.path.join(V3, "reports", "OPUS5_LABEL_QUALITY_REPORT.md")).read()
    assert "contributed evidence to the provisional reference" in text
    assert "descriptive" in text
    assert "human-final labels" in text


# ---------------------------------------------------------------- uncertainty ------------

@pytest.mark.parametrize("ss", SETS)
def test_uncertain_is_preserved_not_collapsed(all_records, ss):
    labs = [r["opus5_provisional_label"] for r in all_records[ss]["records"]]
    assert "UNCERTAIN" in labs, "UNCERTAIN must remain a first-class class, never forced binary"


def test_binary_evaluations_exclude_uncertain_and_report_coverage():
    gd = json.load(open(os.path.join(OPUS5, "gold_dev_baseline", "gold_dev_baseline_report.json")))
    assert gd["n_uncertain_excluded"] > 0
    assert gd["n_evaluated_binary"] + gd["n_uncertain_excluded"] == gd["n_total_gold_dev_items"]
    assert "uncertain_coverage_pct" in gd


def test_no_unsafe_rests_on_weak_support_only():
    """Instruction: items supported only by SOURCE_RULE_ONLY_SUPPORT or
    INCOMPLETE_GUARD_EVIDENCE should normally be UNCERTAIN, not UNSAFE."""
    weak = {"SOURCE_RULE_ONLY_SUPPORT", "INCOMPLETE_GUARD_EVIDENCE"}
    for ss in SETS:
        for r in load(ss)["records"]:
            if r["opus5_provisional_label"] == "UNSAFE":
                assert r["unsafe_support_class"] not in weak, r["item_id"]


def test_every_safe_or_unsafe_cites_concrete_evidence():
    for ss in SETS:
        for r in load(ss)["records"]:
            lab = r["opus5_provisional_label"]
            if lab == "UNSAFE":
                assert r["concrete_unsafe_paths"] != "none identified", r["item_id"]
            elif lab == "SAFE":
                cites = (r["concrete_safe_controls"] != "none identified"
                         or "analysed to completion" in r["final_rationale"])
                assert cites, r["item_id"]


# ---------------------------------------------------------------- blinding ---------------

FORBIDDEN_MODEL_FIELDS = ("ref_model_mean_score", "authguard_score", "authguard_prediction",
                          "is_false_positive", "is_false_negative", "gold_dev_stratum",
                          "calibrated_score", "model_score")


@pytest.mark.parametrize("ss", SETS)
def test_dossiers_contain_no_model_output(ss):
    text = open(os.path.join(OPUS5, "dossiers", f"{ss}_dossiers.json")).read()
    for field in FORBIDDEN_MODEL_FIELDS:
        assert field not in text, f"{field} leaked into {ss} dossiers"


@pytest.mark.parametrize("ss", SETS)
def test_labels_contain_no_model_output(all_records, ss):
    text = json.dumps(all_records[ss])
    for field in FORBIDDEN_MODEL_FIELDS:
        assert field not in text


def test_dossier_builder_never_reads_model_columns():
    src = open(os.path.join(V3, "experiments", "opus5_labeling", "build_dossiers.py")).read()
    for field in ("ref_model_mean_score", "gold_dev_stratum"):
        # may appear only in the forbidden-list declaration, never in a read
        assert f'row["{field}"]' not in src
        assert f"row.get('{field}')" not in src


def test_static_analyzer_evidence_is_deliberately_present():
    """This pass is defined by NOT blinding the labeler from the analyzer."""
    for ss in SETS:
        d = json.load(open(os.path.join(OPUS5, "dossiers", f"{ss}_dossiers.json")))
        for x in d["dossiers"]:
            s = x["source_static_analyzer_evidence"]
            assert s["source_rule_label"] in {"positive", "unflagged"}
            assert "rule_firing_tuples_for_this_address" in s
            assert s["rule_models_authorization"] is False


# ---------------------------------------------------------------- gold-test hygiene -------

def test_model_selection_manifest_never_references_gold_test():
    p = os.path.join(OPUS5, "provisional_final_model_manifest.json")
    text = open(p).read()
    assert "gold_test" not in text.lower(), "Gold-Test must not participate in model selection"


def test_cascade_policy_was_developed_on_gold_dev_only():
    d = json.load(open(os.path.join(OPUS5, "cascade", "cascade_report.json")))
    text = json.dumps(d)
    assert "gold_dev" in text
    band = d.get("escalation_band_selected_on_gold_dev_only")
    assert band, "the escalation band must be recorded so its provenance is auditable"


def test_gold_dev_and_gold_test_families_are_disjoint():
    def fams(name):
        with open(os.path.join(HUMAN_EVAL, name), newline="") as f:
            return {r["family_id"] for r in csv.DictReader(f)}
    assert not (fams("gold_dev_manifest.csv") & fams("gold_test_manifest.csv"))


@pytest.mark.parametrize("name,expected", sorted(MANIFEST_MD5.items()))
def test_sampling_manifests_are_unmodified(name, expected):
    p = os.path.join(HUMAN_EVAL, name)
    assert hashlib.md5(open(p, "rb").read()).hexdigest() == expected, f"{name} was modified"


# ---------------------------------------------------------------- output separation -------

def test_previous_provisional_results_are_not_overwritten():
    """The previous pass's labels must still be readable and still say LLM_PROVISIONAL."""
    for ss in SETS:
        d = json.load(open(os.path.join(PREV, f"{ss}_labels.json")))
        assert d["LABEL_SOURCE"] == "LLM_PROVISIONAL"
    assert os.path.isdir(OPUS5) and os.path.isdir(PREV)
    assert os.path.abspath(OPUS5) != os.path.abspath(PREV)


@pytest.mark.parametrize("ss", SETS)
def test_pipeline_projection_exactly_matches_opus5_reviews(ss):
    reviews = {
        row["item_id"]: row["opus5_provisional_label"]
        for row in load(ss)["records"]
    }
    projection = json.load(open(os.path.join(OPUS5, f"{ss}_labels.json")))
    projected = {
        row["item_id"]: row["llm_provisional_label"]
        for row in projection["records"]
    }
    assert projected == reviews


def test_opus5_results_live_only_in_their_own_directory():
    for sub in ("gold_dev_baseline", "gold_test", "cascade", "retraining"):
        assert os.path.isdir(os.path.join(OPUS5, sub)), sub
    assets = os.path.join(V3, "manuscript_assets", "provisional_llm_provisional_opus5")
    assert os.path.isdir(assets)
    text = open(os.path.join(assets, "table_08_provisional_gold_test_results.md")).read()
    assert "OPUS 5 LABELS WITH STATIC-ANALYZER EVIDENCE" in text


def test_phase2_frozen_model_config_untouched():
    p = os.path.join(V3, "configs", "final_model.json")
    d = json.load(open(p))
    assert "provisional" not in json.dumps(d).lower() or d.get("model_name") != "provisional_final_model"


def test_previous_provisional_config_not_replaced():
    prev = json.load(open(os.path.join(V3, "configs", "provisional_final_model.json")))
    assert prev["LABEL_SOURCE"] == "LLM_PROVISIONAL"
    opus = os.path.join(V3, "configs", "provisional_final_model_llm_provisional_opus5.json")
    assert os.path.exists(opus)
    assert json.load(open(opus))["LABEL_SOURCE"] == "LLM_PROVISIONAL_OPUS5"


# ---------------------------------------------------------------- analysis honesty --------

@pytest.mark.parametrize("ss", SETS)
def test_coverage_gaps_are_recorded_not_hidden(ss):
    d = json.load(open(os.path.join(OPUS5, "dossiers", f"{ss}_dossiers.json")))
    for x in d["dossiers"]:
        cfg = x["cfg_guard_analysis_opus5"]
        if "error" in cfg:
            continue
        assert "static_opcode_census" in cfg
        assert "sensitive_opcodes_never_reached_by_analysis" in cfg


def test_items_with_a_coverage_gap_never_claim_no_dangerous_path():
    """A SAFE/NO_CONCRETE_DANGEROUS_PATH_FOUND claim is not allowed while the analysis is
    known to have missed sensitive instructions."""
    for ss in SETS:
        doss = {x["item_id"]: x for x in json.load(
            open(os.path.join(OPUS5, "dossiers", f"{ss}_dossiers.json")))["dossiers"]}
        for r in load(ss)["records"]:
            if r["opus5_provisional_label"] != "SAFE":
                continue
            if r["manual_override_applied"] == "YES":
                continue  # override carries its own recorded justification
            cfg = doss[r["item_id"]]["cfg_guard_analysis_opus5"]
            assert not cfg.get("sensitive_opcodes_never_reached_by_analysis"), r["item_id"]


def test_manual_overrides_all_carry_a_written_justification():
    from overrides import OVERRIDES
    for iid, ov in OVERRIDES.items():
        assert ov.get("reason_text"), iid
        assert len(ov["reason_text"]) > 80, f"{iid}: override justification is too thin"
        assert ov["label"] in LABELS


def test_override_records_are_marked_in_the_output():
    from overrides import OVERRIDES
    seen = set()
    for ss in SETS:
        for r in load(ss)["records"]:
            if r["item_id"] in OVERRIDES:
                seen.add(r["item_id"])
                assert r["manual_override_applied"] == "YES"
                assert "MANUAL REVIEW OVERRIDE" in r["final_rationale"]
    assert seen == set(OVERRIDES)


# ---------------------------------------------------------------- reproducibility ---------

def test_one_command_rerun_supports_the_new_label_source():
    src = open(os.path.join(V3, "run_reference_pipeline.py")).read()
    assert "llm_provisional_opus5" in src
    assert "AUTHGUARD_LABEL_SOURCE" in src


def test_human_final_mode_still_refuses_to_fabricate():
    r = subprocess.run(
        [sys.executable, os.path.join(V3, "run_reference_pipeline.py"),
         "--label-source", "human_final"],
        capture_output=True, text=True, cwd=ROOT, timeout=600)
    assert r.returncode == 0
    assert "BLOCKED" in r.stdout
    manifest = os.path.join(V3, "results", "human_final", "run_manifest.json")
    assert json.load(open(manifest))["status"] == "BLOCKED_NO_HUMAN_LABELS"


def test_pipeline_scripts_default_to_the_previous_label_source():
    """Unset AUTHGUARD_LABEL_SOURCE must reproduce the previous pass's paths exactly."""
    for name in ("run_gold_dev_baseline.py", "run_gold_test_evaluation.py",
                 "run_cascade_evaluation.py", "run_retraining_experiments.py",
                 "select_provisional_final_model.py"):
        src = open(os.path.join(V3, "experiments", "llm_provisional", name)).read()
        assert 'os.environ.get("AUTHGUARD_LABEL_SOURCE", "llm_provisional")' in src, name


def test_workbooks_gained_the_opus5_columns_and_kept_human_columns_blank():
    from openpyxl import load_workbook
    manifest = json.load(open(os.path.join(OPUS5, "workbook_update_manifest.json")))
    for res in manifest["results"]:
        assert res["status"] == "UPDATED", res
    for wb_name in ("Pilot_Code_Review.xlsx", "Gold_Dev_Code_Review.xlsx",
                    "Gold_Test_Code_Review.xlsx"):
        wb = load_workbook(os.path.join(HUMAN_EVAL, wb_name), read_only=True)
        ws = wb["REVIEW_ITEMS"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for col in ("opus5_proposed_label", "opus5_confidence", "opus5_rationale",
                    "source_static_analyzer_verdict", "guard_tracer_result",
                    "opus5_unresolved_questions"):
            assert col in headers, f"{wb_name} missing {col}"
        idx = {h: i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=2, values_only=True):
            for col in ("contributor_label", "final_label", "final_rationale"):
                if col in idx:
                    assert row[idx[col]] in (None, ""), f"{wb_name}: {col} must stay blank"
        wb.close()
