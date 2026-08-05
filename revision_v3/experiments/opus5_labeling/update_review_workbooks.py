"""Add the Opus 5 review columns to the three human-review workbooks.

Adds, per item: the Opus 5 proposed label, confidence and rationale; the source static-analyzer
verdict and a plain explanation of what that verdict does and does not mean; a readable code
snippet; the guard-tracer result; project / on-chain evidence; and the unresolved questions.

Human columns (`contributor_label`, `contributor_rationale`, `group_discussion`,
`final_label`, `final_rationale`) are left exactly as they are — and the script refuses to
touch a workbook if any of them is already filled, so in-flight review can never be
overwritten. A timestamped backup of each workbook is written before any modification.

Reviewers are never required to read raw bytecode: the snippet column is a rendered summary of
what the analysis found, and every claim is stated in words.
"""

from __future__ import annotations

import json
import os
import shutil
from datetime import datetime, timezone

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
HUMAN_EVAL = os.path.join(ROOT, "human_eval")
OUT = os.path.join(ROOT, "results", "llm_provisional_opus5")

WORKBOOKS = {
    "pilot": "Pilot_Code_Review.xlsx",
    "gold_dev": "Gold_Dev_Code_Review.xlsx",
    "gold_test": "Gold_Test_Code_Review.xlsx",
}
HUMAN_COLUMNS = ["contributor_label", "contributor_rationale", "group_discussion",
                 "final_label", "final_rationale"]

NEW_COLUMNS = [
    ("opus5_proposed_label", 16),
    ("opus5_confidence", 13),
    ("opus5_rationale", 80),
    ("source_static_analyzer_verdict", 18),
    ("source_static_analyzer_explanation", 80),
    ("opus5_relevant_code_summary", 80),
    ("guard_tracer_result", 60),
    ("project_and_onchain_evidence", 50),
    ("opus5_unresolved_questions", 60),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
NEW_FILL = PatternFill("solid", fgColor="EAF1FA")


def guard_summary(dossier: dict) -> str:
    cfg = dossier["cfg_guard_analysis_opus5"]
    if "error" in cfg:
        return cfg["error"]
    lines = []
    for fn in cfg.get("per_function", []):
        name = fn.get("resolved_signature") or fn["selector"]
        status = {
            "UNGUARDED_PATH": "an unauthenticated path to a sensitive operation exists",
            "GUARD_DOMINATED": "every path to a sensitive operation passes an authorization check",
            "GUARDED_BY_STORAGE_CONDITION": "gated only by a storage-dependent condition",
            "NO_SENSITIVE_OP": "no sensitive operation reachable",
        }.get(fn["guard_status"], fn["guard_status"])
        extra = ""
        if fn.get("guards"):
            extra = " — check found: " + "; ".join(
                g["semantics"].split(" (")[0]
                + (f" against {g['compared_address_constant']}"
                   if g.get("compared_address_constant") else "")
                for g in fn["guards"][:2])
        if fn.get("analysis_incomplete"):
            extra += " [analysis incomplete for this function]"
        lines.append(f"• {name}: {status}{extra}")
    gap = cfg.get("sensitive_opcodes_never_reached_by_analysis")
    if gap:
        lines.append(f"• COVERAGE GAP: the analysis never reached {sum(len(v) for v in gap.values())} "
                     f"sensitive instruction(s) that exist in this contract, so the list above is "
                     f"a lower bound on what it can do.")
    return "\n".join(lines) or "no dispatched functions recovered"


def code_summary(dossier: dict, rec: dict) -> str:
    """A words-first summary of what the code does — no raw bytecode required."""
    cfg = dossier["cfg_guard_analysis_opus5"]
    parts = [rec["contract_purpose"]]
    if rec["concrete_unsafe_paths"] != "none identified":
        parts.append("Concern(s) found: " + rec["concrete_unsafe_paths"])
    if rec["concrete_safe_controls"] != "none identified":
        parts.append("Protection(s) found: " + rec["concrete_safe_controls"])
    census = cfg.get("static_opcode_census") or {}
    if census:
        parts.append("Instruction census (whole contract): "
                     + ", ".join(f"{k}×{v}" for k, v in sorted(census.items())))
    return "\n\n".join(parts)


def project_evidence(dossier: dict) -> str:
    ident = dossier["identity"]
    code = dossier["source_and_code_evidence"]
    bits = [f"chain={ident['chain']}", f"runtime size={ident['runtime_size_bytes']} bytes",
            f"verified source={code.get('verified_source')}",
            f"other chains with identical bytecode: {ident.get('chains_sharing_this_exact_bytecode')}"]
    proj = ident.get("documented_project")
    if proj:
        bits.append(f"documented project: {proj.get('project')} "
                    f"({proj.get('category')}, provenance {proj.get('provenance_confidence')}); "
                    f"docs: {proj.get('official_documentation')}")
    live = code.get("live_storage_reads")
    if live:
        bits.append(f"on-chain storage read during evidence collection: {json.dumps(live)[:300]}")
    consts = code.get("embedded_address_constants")
    if consts:
        bits.append(f"addresses embedded in the code: {json.dumps(consts)[:300]}")
    return "\n".join(bits)


ANALYZER_EXPLANATION = {
    "positive": (
        "The original static analyzer FLAGGED this contract. Its rule is a single reachability "
        "test: 'an external call is reachable from receive() or fallback()'. It contains no "
        "authorization check at all, so a flag means a powerful operation can be reached from "
        "an entry point that carries no caller check at dispatch time — it does NOT by itself "
        "mean access control is missing."),
    "unflagged": (
        "The original static analyzer did NOT flag this contract. That is a weak signal: it "
        "only means this one reachability pattern did not fire. The dataset's negatives are "
        "rule-silent delegates with no benignity verification of any kind, so 'unflagged' is "
        "not evidence of safety."),
}


def update(sample_set: str) -> dict:
    path = os.path.join(HUMAN_EVAL, WORKBOOKS[sample_set])
    if not os.path.exists(path):
        return {"sample_set": sample_set, "status": "WORKBOOK_NOT_FOUND", "path": path}

    recs = {r["item_id"]: r for r in
            json.load(open(os.path.join(OUT, f"{sample_set}_reviews_opus5.json")))["records"]}
    doss = {x["item_id"]: x for x in
            json.load(open(os.path.join(OUT, "dossiers", f"{sample_set}_dossiers.json")))["dossiers"]}

    wb = load_workbook(path)
    ws = wb["REVIEW_ITEMS"] if "REVIEW_ITEMS" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    # Refuse to touch a workbook whose human columns already carry review work.
    filled = []
    for col in HUMAN_COLUMNS:
        if col not in idx:
            continue
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=idx[col]).value
            if v not in (None, ""):
                filled.append((col, row))
    if filled:
        return {"sample_set": sample_set, "status": "REFUSED_HUMAN_WORK_PRESENT",
                "n_filled_human_cells": len(filled), "examples": filled[:5]}

    backup = path.replace(".xlsx", f".backup-{datetime.now(timezone.utc):%Y%m%dT%H%M%SZ}.xlsx")
    shutil.copy2(path, backup)

    item_col = idx.get("item_id") or 1
    start = ws.max_column + 1
    for offset, (name, width) in enumerate(NEW_COLUMNS):
        col = start + offset
        c = ws.cell(row=1, column=col, value=name)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    n = 0
    for row in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row, column=item_col).value
        r, d = recs.get(item_id), doss.get(item_id)
        if not r or not d:
            continue
        values = [
            r["opus5_provisional_label"], r["opus5_confidence"], r["final_rationale"],
            f'{r["source_rule_label"]} — assessed {r["static_analyzer_verdict_assessment"]}',
            ANALYZER_EXPLANATION[r["source_rule_label"]] + "\n\nAssessment against the "
            "control-flow evidence for this item: " + r["source_rule_assessment"],
            code_summary(d, r), guard_summary(d), project_evidence(d),
            r["unresolved_questions"],
        ]
        for offset, v in enumerate(values):
            c = ws.cell(row=row, column=start + offset, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.fill = NEW_FILL
        n += 1

    wb.save(path)
    return {"sample_set": sample_set, "status": "UPDATED", "rows_updated": n,
            "columns_added": [c for c, _ in NEW_COLUMNS], "backup": os.path.basename(backup),
            "human_columns_left_blank": HUMAN_COLUMNS}


def main() -> int:
    out = [update(ss) for ss in WORKBOOKS]
    path = os.path.join(OUT, "workbook_update_manifest.json")
    with open(path, "w") as f:
        json.dump({"LABEL_SOURCE": "LLM_PROVISIONAL_OPUS5",
                   "STATIC_ANALYZER_EVIDENCE": "VISIBLE",
                   "STATUS": "PROVISIONAL_PENDING_HUMAN_REVIEW",
                   "generated_at_utc": datetime.now(timezone.utc).isoformat(),
                   "results": out}, f, indent=1)
    for o in out:
        print(o["sample_set"], o["status"], o.get("rows_updated", ""))
    print("wrote", path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
