"""Builds revision_v3/human_eval/Pilot_Master_Adjudication.xlsx: evidence + LLM review +
(initially empty) contributor-response sections + disagreement summary + discussion notes +
lead-author final columns. Contributor sections are appended later, one per reviewer, by
import_reviewer_workbook.py -- this script only creates the starting structure.
"""
from __future__ import annotations

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from excel_builder import (  # noqa: E402
    EVIDENCE_COLUMNS, HEADER_FILL, HEADER_FONT, LEAD_AUTHOR_FILL, LLM_COLUMNS, LLM_FILL,
    build_examples_sheet, build_guide_sheet, build_start_here_sheet, load_llm_reviews,
    load_manifest_rows, load_packets_by_item_id,
)
from excel_builder import _evidence_row_from_manifest_row  # noqa: E402

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
HUMAN_EVAL_DIR = os.path.join(REPO_ROOT, "revision_v3", "human_eval")

MASTER_META_COLUMNS = ["disagreement_summary", "discussion_notes_combined"]
LEAD_AUTHOR_COLUMNS_MASTER = ["final_label", "final_reason_category", "final_rationale", "final_decision_date"]

ITEMS_SHEET = "MASTER_ITEMS"


def build_master_workbook(manifest_path: str, evidence_dump_path: str, llm_reviews_path: str,
                           output_path: str, review_set_name: str = "Pilot") -> None:
    manifest_rows = load_manifest_rows(manifest_path)
    packets_by_item = load_packets_by_item_id(evidence_dump_path, manifest_rows)
    llm_reviews = load_llm_reviews(llm_reviews_path)

    wb = Workbook()
    wb.remove(wb.active)

    build_start_here_sheet(wb, review_set_name=f"{review_set_name} MASTER ADJUDICATION", n_items=len(manifest_rows))
    build_guide_sheet(wb)
    build_examples_sheet(wb)

    ws = wb.create_sheet(ITEMS_SHEET)
    base_columns = EVIDENCE_COLUMNS + LLM_COLUMNS + MASTER_META_COLUMNS + LEAD_AUTHOR_COLUMNS_MASTER

    for col_idx, col_name in enumerate(base_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        width = 40 if col_name not in ("item_id", "chain", "contract_address") else 24
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(manifest_rows, start=2):
        packet = packets_by_item[row["item_id"]]
        evidence = _evidence_row_from_manifest_row(row, packet)
        llm = llm_reviews.get(row["item_id"], {})
        record = dict(evidence)
        for c in LLM_COLUMNS:
            record[c] = llm.get(c, "")
        record["disagreement_summary"] = "(no contributor responses imported yet)"
        record["discussion_notes_combined"] = ""
        for c in LEAD_AUTHOR_COLUMNS_MASTER:
            record[c] = ""

        for col_idx, col_name in enumerate(base_columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_name in LLM_COLUMNS:
                cell.fill = LLM_FILL
            elif col_name in LEAD_AUTHOR_COLUMNS_MASTER:
                cell.fill = LEAD_AUTHOR_FILL

    ws.freeze_panes = ws.cell(row=2, column=len(EVIDENCE_COLUMNS) + 1).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(len(base_columns))}{len(manifest_rows) + 1}"
    ws.row_dimensions[1].height = 30

    # metadata sheet tracking which reviewers have been imported (used by the importer for
    # duplicate-reviewer detection)
    meta = wb.create_sheet("_IMPORT_LOG")
    meta.append(["reviewer_name", "source_file", "imported_at_utc", "n_items_imported"])
    for c in meta[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    wb.active = 0
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"wrote {output_path} ({len(manifest_rows)} items, 0 contributor sections yet)")


def main() -> int:
    build_master_workbook(
        manifest_path=os.path.join(HUMAN_EVAL_DIR, "pilot_manifest.csv"),
        evidence_dump_path=os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_evidence_dump.json"),
        llm_reviews_path=os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_llm_reviews.json"),
        output_path=os.path.join(HUMAN_EVAL_DIR, "Pilot_Master_Adjudication.xlsx"),
        review_set_name="Pilot",
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
