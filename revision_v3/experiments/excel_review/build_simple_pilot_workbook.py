"""Builds Pilot_Simple_Review.xlsx -- the single simplified workbook sent to Pilot contributors.

This is a SIMPLIFICATION of the Pilot review process, not a replacement of Phase 3A's
infrastructure: it reuses the existing, unmodified Phase 3A inputs (the frozen 20-item Pilot
manifest, the evidence dump, and the technical LLM reviews) purely as read-only sources. It
does not resample, modify, or import from Pilot_Review.xlsx, Pilot_Master_Adjudication.xlsx,
create_reviewer_copy.py, or import_reviewer_workbook.py -- those remain in place, unused by
this workflow, as archived supporting infrastructure.

There is no reviewer-copy step and no import step for this workflow: contributors, the group
discussion, and the lead author all work directly in this one file.

Usage:
    python3 revision_v3/experiments/excel_review/build_simple_pilot_workbook.py
"""
from __future__ import annotations

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from excel_builder import load_manifest_rows, load_packets_by_item_id  # noqa: E402

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
HUMAN_EVAL_DIR = os.path.join(REPO_ROOT, "revision_v3", "human_eval")
sys.path.insert(0, HUMAN_EVAL_DIR)
from taxonomy import PRIMARY_LABELS  # noqa: E402

MANIFEST_PATH = os.path.join(HUMAN_EVAL_DIR, "pilot_manifest.csv")
EVIDENCE_DUMP_PATH = os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_evidence_dump.json")
LLM_REVIEWS_PATH = os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_llm_reviews.json")
PLAIN_ENGLISH_PATH = os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_plain_english_reviews.json")
OUTPUT_PATH = os.path.join(HUMAN_EVAL_DIR, "Pilot_Simple_Review.xlsx")

EVIDENCE_COLUMNS = [
    "item_id", "chain", "contract_address", "project_name_if_known",
    "explorer_or_source_link", "verified_source_available", "implementation_resolved",
    "plain_english_contract_summary", "sensitive_actions", "access_control_summary",
    "initialization_summary", "proxy_or_upgrade_summary", "concrete_security_concern",
    "missing_information",
]
AI_COLUMNS = ["ai_proposed_label", "ai_confidence", "ai_explanation", "ai_points_to_check"]
CONTRIBUTOR_COLUMNS = [
    "contributor_1_label", "contributor_1_reason",
    "contributor_2_label", "contributor_2_reason",
    "contributor_3_label", "contributor_3_reason",
]
GROUP_COLUMNS = ["group_discussion_note", "agreed_group_label"]
LEAD_AUTHOR_COLUMNS = ["final_label", "final_reason"]

REVIEW_COLUMNS = EVIDENCE_COLUMNS + AI_COLUMNS + CONTRIBUTOR_COLUMNS + GROUP_COLUMNS + LEAD_AUTHOR_COLUMNS

LABEL_DROPDOWN_COLUMNS = [
    "contributor_1_label", "contributor_2_label", "contributor_3_label",
    "agreed_group_label", "final_label",
]

WIDE_COLUMNS = {
    "plain_english_contract_summary", "sensitive_actions", "access_control_summary",
    "initialization_summary", "proxy_or_upgrade_summary", "concrete_security_concern",
    "missing_information", "ai_explanation", "ai_points_to_check",
    "contributor_1_reason", "contributor_2_reason", "contributor_3_reason",
    "group_discussion_note", "final_reason",
}

TECHNICAL_COLUMNS = [
    "item_id", "runtime_bytecode_hash", "runtime_size", "selector_summary", "opcode_summary",
    "proxy_detection_details", "source_code_link", "decompiler_output_path",
    "documentation_link", "evidence_limitations",
]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ROW_FILL_EVEN = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
LEAD_AUTHOR_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _verified_source_text(packet: dict) -> str:
    status = packet.get("verified_source_code_availability", {}).get("status", "")
    if status == "NOT_DETERMINABLE_OFFLINE":
        return "Unknown (not checked in this offline review)"
    return status or "Unknown"


def _implementation_resolved_text(packet: dict) -> str:
    proxy = packet.get("proxy_evidence", {})
    if proxy.get("is_eip7702_designator") or proxy.get("has_delegatecall"):
        target = proxy.get("designator_target_address")
        if target:
            return f"Yes -- {target}"
        return "No -- forwards to another contract, but that contract could not be identified offline"
    return "Not applicable (not a proxy/forwarder)"


def _project_name_text(packet: dict) -> str:
    known = packet.get("known_project")
    return known["project"] if known else "(none documented)"


def _documentation_link_text(packet: dict) -> str:
    known = packet.get("known_project")
    return known["documentation_url"] if known else "(none documented)"


def _selector_summary_text(packet: dict) -> str:
    proxy = packet.get("proxy_evidence", {})
    token = packet.get("token_transfer_evidence", {})
    admin = [k for k, v in proxy.get("admin_ownership_selectors_present", {}).items() if v]
    moved = [k for k, v in token.get("token_movement_selectors_present", {}).items() if v]
    parts = []
    if admin:
        parts.append("admin/ownership: " + ", ".join(admin))
    if moved:
        parts.append("token-movement: " + ", ".join(moved))
    if token.get("approval_selector_present"):
        parts.append("approval selector present")
    return "; ".join(parts) if parts else "no named admin/ownership/token selectors detected"


def _proxy_detection_text(packet: dict) -> str:
    proxy = packet.get("proxy_evidence", {})
    bits = [
        f"is_eip7702_designator={proxy.get('is_eip7702_designator')}",
        f"has_delegatecall={proxy.get('has_delegatecall')} (count={proxy.get('delegatecall_count')})",
        f"eip1967_implementation_slot_present={proxy.get('eip1967_implementation_slot_present')}",
        f"resembles_minimal_forwarder={proxy.get('resembles_minimal_forwarder')}",
    ]
    return "; ".join(bits)


def build_evidence_row(item_id: str, chain: str, address: str, packet: dict, plain: dict, llm: dict) -> dict:
    ai_explanation = (
        f"{plain['plain_english_contract_summary']} "
        f"Sensitive actions: {plain['sensitive_actions']} "
        f"Access control: {plain['access_control_summary']} "
        f"Initialization: {plain['initialization_summary']} "
        f"Proxy/upgrade: {plain['proxy_or_upgrade_summary']} "
        f"Proposed label: {llm.get('llm_proposed_label', '')} "
        f"({llm.get('llm_confidence', '')} confidence) -- {plain['concrete_security_concern']}"
    )
    return {
        "item_id": item_id,
        "chain": chain,
        "contract_address": address,
        "project_name_if_known": _project_name_text(packet),
        "explorer_or_source_link": packet.get("explorer_link", ""),
        "verified_source_available": _verified_source_text(packet),
        "implementation_resolved": _implementation_resolved_text(packet),
        "plain_english_contract_summary": plain["plain_english_contract_summary"],
        "sensitive_actions": plain["sensitive_actions"],
        "access_control_summary": plain["access_control_summary"],
        "initialization_summary": plain["initialization_summary"],
        "proxy_or_upgrade_summary": plain["proxy_or_upgrade_summary"],
        "concrete_security_concern": plain["concrete_security_concern"],
        "missing_information": plain["missing_information"],
        "ai_proposed_label": llm.get("llm_proposed_label", ""),
        "ai_confidence": llm.get("llm_confidence", ""),
        "ai_explanation": ai_explanation,
        "ai_points_to_check": plain["ai_points_to_check"],
    }


def build_technical_row(item_id: str, packet: dict) -> dict:
    return {
        "item_id": item_id,
        "runtime_bytecode_hash": packet.get("runtime_bytecode_sha256", ""),
        "runtime_size": packet.get("runtime_bytecode_length_bytes", ""),
        "selector_summary": _selector_summary_text(packet),
        "opcode_summary": packet.get("deterministic_summary", ""),
        "proxy_detection_details": _proxy_detection_text(packet),
        "source_code_link": packet.get("explorer_link", ""),
        "decompiler_output_path": "NOT_AVAILABLE_OFFLINE (no decompiler wired into this pipeline)",
        "documentation_link": _documentation_link_text(packet),
        "evidence_limitations": (
            "No decompiler output (only a 60-opcode disassembly prefix plus aggregate "
            "structural counts is available). No live verified-source check performed. No "
            "on-chain authorization or transaction history available in this evidence packet."
        ),
    }


def build_read_me_sheet(wb: Workbook, n_items: int) -> None:
    ws = wb.create_sheet("READ_ME", 0)
    ws.column_dimensions["A"].width = 115
    lines: list[tuple[str, str]] = [
        ("AuthGuard-7702 Pilot Review — READ ME (start here)", "title"),
        ("", ""),
        ("What is EIP-7702?", "h2"),
        ("1. EIP-7702 lets a normal wallet account (an EOA) temporarily run code from a "
         "separate 'delegate' contract, as if that code were its own.", ""),
        ("2. While delegated, that code can act using the wallet's own assets, token "
         "approvals, storage, and identity -- it is not sandboxed away from the wallet.", ""),
        ("3. The main security question for every item in this review is: could someone "
         "other than the wallet owner misuse the delegate contract to do something the "
         "owner did not intend?", ""),
        ("4. You do not need to inspect raw bytecode. Every item below has already been "
         "summarized in plain English, with an AI-generated preliminary analysis attached.", ""),
        ("5. The AI analysis is a starting point, not the final answer -- it can be wrong, "
         "and it is deliberately cautious (it says UNCERTAIN whenever the evidence is "
         "incomplete, rather than guessing).", ""),
        ("6. You may discuss difficult items with other contributors before recording your "
         "answer -- this is a collaborative review, not an independent/blinded one.", ""),
        ("7. The lead author reads every contributor's answer and every discussion note, "
         "and then makes the final decision. Only that final decision is used later.", ""),
        ("", ""),
        ("Four questions to ask about every item", "h2"),
        ("A. Who can trigger sensitive actions? (Anyone? Only the wallet owner? Unclear?)", ""),
        ("B. Can the contract transfer assets, approve tokens, execute arbitrary calls, "
         "change ownership, or upgrade its own code?", ""),
        ("C. Is initialization / setup protected against being run more than once or by the "
         "wrong party?", ""),
        ("D. Is the real implementation visible and understandable, or does this contract "
         "forward to another contract we cannot see?", ""),
        ("", ""),
        ("The three labels", "h2"),
        ("SAFE — the available evidence shows that sensitive actions appear properly "
         "restricted, and no concrete authorization-related vulnerability was identified.", ""),
        ("UNSAFE — the evidence identifies a concrete dangerous condition, such as "
         "unrestricted asset movement, arbitrary execution, unsafe initialization, ownership "
         "takeover, or an unprotected upgrade path.", ""),
        ("UNCERTAIN — the available evidence is not sufficient to make a reliable SAFE or "
         "UNSAFE decision. There is no penalty for choosing UNCERTAIN -- it is often the "
         "correct, honest answer.", ""),
        ("", ""),
        ("How to work through this workbook", "h2"),
        ("1. Read this sheet (about 5 minutes).", ""),
        ("2. Open REVIEW_ITEMS. Read the AI explanation for an item.", ""),
        ("3. Check the plain-English evidence summary next to it.", ""),
        ("4. Open the source/explorer link only if you need more context.", ""),
        ("5. Select SAFE, UNSAFE, or UNCERTAIN in your contributor column.", ""),
        ("6. Write one or two sentences explaining your decision.", ""),
        ("7. Discuss difficult cases with the other contributors.", ""),
        ("8. Record the group's conclusion in group_discussion_note / agreed_group_label.", ""),
        ("9. The lead author records the final_label and final_reason (those two columns "
         "are locked -- contributors cannot edit them).", ""),
        ("", ""),
        ("Suggested first meeting (calibration)", "h2"),
        (f"Review the first 5 of the {n_items} Pilot items together in one meeting: briefly "
         "explain EIP-7702, walk through one SAFE example, one UNSAFE example, and one "
         "UNCERTAIN example (see the worked examples referenced in REVIEWER_GUIDE.md / "
         "Pilot_Review.xlsx's EXAMPLES sheet), then discuss those 5 real items together. "
         "After that, the remaining items may be reviewed individually or together.", ""),
        ("", ""),
        ("Optional technical sheet", "h2"),
        ("A third sheet, TECHNICAL_DETAILS, contains bytecode hashes, selector lists, and "
         "opcode-level summaries. You do not need it unless you have smart-contract "
         "experience and want to dig deeper on a specific item.", ""),
    ]
    row = 1
    for text, style in lines:
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if style == "title":
            cell.font = Font(bold=True, size=16)
        elif style == "h2":
            cell.font = Font(bold=True, size=12)
        row += 1
    ws.freeze_panes = "A2"


def build_review_items_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("REVIEW_ITEMS")

    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        width = 45 if col_name in WIDE_COLUMNS else 22
        if col_name in ("item_id", "explorer_or_source_link"):
            width = 30
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    for row_idx, record in enumerate(rows, start=2):
        row_fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
        for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = LEAD_AUTHOR_FILL if col_name in LEAD_AUTHOR_COLUMNS else row_fill

    last_row = len(rows) + 1
    last_col = len(REVIEW_COLUMNS)

    def col_letter(name: str) -> str:
        return get_column_letter(REVIEW_COLUMNS.index(name) + 1)

    for col_name in LABEL_DROPDOWN_COLUMNS:
        dv = DataValidation(type="list", formula1=f'"{",".join(PRIMARY_LABELS)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter(col_name)}2:{col_letter(col_name)}{last_row}")

    ws.freeze_panes = ws.cell(row=2, column=2).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


def protect_lead_author_columns(ws, n_rows: int) -> None:
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        locked = col_name in LEAD_AUTHOR_COLUMNS
        for row in range(1, n_rows + 2):
            ws.cell(row=row, column=col_idx).protection = Protection(locked=locked)
    ws.protection = SheetProtection(
        sheet=True, password=None,
        formatCells=False, formatColumns=False, formatRows=False,
        insertRows=False, insertColumns=False, deleteRows=False, deleteColumns=False,
        sort=False, autoFilter=False, pivotTables=True,
        selectLockedCells=False, selectUnlockedCells=False,
    )


def build_technical_details_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("TECHNICAL_DETAILS")
    ws.merge_cells("A1:J1")
    banner = ws.cell(row=1, column=1, value=(
        "OPTIONAL TECHNICAL REFERENCE — contributors are not expected to review this sheet "
        "unless they have the necessary technical knowledge."
    ))
    banner.font = Font(bold=True, color="C00000", size=12)
    banner.fill = WARN_FILL
    banner.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30

    for col_idx, col_name in enumerate(TECHNICAL_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        width = 45 if col_name in ("opcode_summary", "proxy_detection_details", "evidence_limitations") else 26
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, record in enumerate(rows, start=3):
        row_fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
        for col_idx, col_name in enumerate(TECHNICAL_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = row_fill

    last_row = len(rows) + 2
    last_col = len(TECHNICAL_COLUMNS)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_row}"


def main() -> int:
    manifest_rows = load_manifest_rows(MANIFEST_PATH)
    assert len(manifest_rows) == 20, f"expected 20 Pilot items, found {len(manifest_rows)}"

    packets_by_item = load_packets_by_item_id(EVIDENCE_DUMP_PATH, manifest_rows)

    with open(LLM_REVIEWS_PATH) as f:
        llm_reviews = json.load(f)["reviews"]
    with open(PLAIN_ENGLISH_PATH) as f:
        plain_reviews = json.load(f)["reviews"]

    evidence_rows = []
    technical_rows = []
    for row in manifest_rows:
        item_id = row["item_id"]
        packet = packets_by_item[item_id]
        plain = plain_reviews[item_id]
        llm = llm_reviews[item_id]
        evidence_rows.append(
            build_evidence_row(item_id, row["chain"], row["address"], packet, plain, llm)
        )
        technical_rows.append(build_technical_row(item_id, packet))

    wb = Workbook()
    wb.remove(wb.active)
    build_read_me_sheet(wb, n_items=len(manifest_rows))
    build_review_items_sheet(wb, evidence_rows)
    protect_lead_author_columns(wb["REVIEW_ITEMS"], n_rows=len(evidence_rows))
    build_technical_details_sheet(wb, technical_rows)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH} ({len(evidence_rows)} items)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
