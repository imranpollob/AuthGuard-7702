"""Builds Pilot_Code_Review.xlsx -- the code-evidence-backed simplified Pilot workbook.

This supersedes Pilot_Simple_Review.xlsx as the file sent to contributors (per explicit
instruction: the previous simplified workbook lacked real code evidence -- only opcode-count
summaries). It does NOT modify Pilot_Simple_Review.xlsx, Pilot_Review.xlsx,
Pilot_Master_Adjudication.xlsx, or any Phase 3A file -- those remain in place, unused, as
archived infrastructure.

Reads (read-only):
  - pilot_manifest.csv (frozen, 20 items -- not resampled)
  - pilot_code_evidence/<item>/verification_status.json (Sourcify + Blockscout results)
  - llm_reviews/pilot_code_review_content.json (per-item narrative fields, grounded in the
    decompiled evidence under pilot_code_evidence/)

Usage:
    python3 revision_v3/experiments/excel_review/build_code_review_workbook.py
"""
from __future__ import annotations

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from excel_builder import load_manifest_rows, load_packets_by_item_id  # noqa: E402

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
HUMAN_EVAL_DIR = os.path.join(REPO_ROOT, "revision_v3", "human_eval")
sys.path.insert(0, HUMAN_EVAL_DIR)
from taxonomy import PRIMARY_LABELS  # noqa: E402

MANIFEST_PATH = os.path.join(HUMAN_EVAL_DIR, "pilot_manifest.csv")
EVIDENCE_DUMP_PATH = os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_evidence_dump.json")
CODE_REVIEW_CONTENT_PATH = os.path.join(HUMAN_EVAL_DIR, "llm_reviews", "pilot_code_review_content.json")
CODE_EVIDENCE_DIR = os.path.join(HUMAN_EVAL_DIR, "pilot_code_evidence")
OUTPUT_PATH = os.path.join(HUMAN_EVAL_DIR, "Pilot_Code_Review.xlsx")

EVIDENCE_COLUMNS = [
    "item_id", "chain", "address", "project_name", "explorer_link",
    "verified_source_status", "implementation_address", "implementation_resolved",
    "contract_purpose", "sensitive_function_names", "relevant_code_snippet",
    "code_source_or_decompiler_reference", "access_control_explanation",
    "initialization_explanation", "proxy_and_upgrade_explanation",
    "asset_transfer_or_approval_explanation", "concrete_security_finding",
    "unresolved_questions",
]
AI_COLUMNS = ["ai_proposed_label", "ai_confidence", "ai_reasoning"]
CONTRIBUTOR_COLUMNS = ["contributor_label", "contributor_reason"]
GROUP_COLUMNS = ["group_discussion"]
LEAD_AUTHOR_COLUMNS = ["final_label", "final_reason"]

REVIEW_COLUMNS = EVIDENCE_COLUMNS + AI_COLUMNS + CONTRIBUTOR_COLUMNS + GROUP_COLUMNS + LEAD_AUTHOR_COLUMNS

LABEL_DROPDOWN_COLUMNS = ["contributor_label", "final_label"]

WIDE_COLUMNS = {
    "contract_purpose", "sensitive_function_names", "relevant_code_snippet",
    "code_source_or_decompiler_reference", "access_control_explanation",
    "initialization_explanation", "proxy_and_upgrade_explanation",
    "asset_transfer_or_approval_explanation", "concrete_security_finding",
    "unresolved_questions", "ai_reasoning", "contributor_reason", "group_discussion",
    "final_reason",
}
EXTRA_WIDE_COLUMNS = {"relevant_code_snippet"}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ROW_FILL_EVEN = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
LEAD_AUTHOR_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
CODE_FONT = Font(name="Consolas", size=9)


def load_verification_status(item_id: str) -> dict:
    folder = os.path.join(CODE_EVIDENCE_DIR, item_id.replace(":", "_"))
    with open(os.path.join(folder, "verification_status.json")) as f:
        return json.load(f)


def verified_source_status_text(verification: dict) -> str:
    if verification.get("verified"):
        return "VERIFIED (source available)"
    return "NOT VERIFIED (checked live: Sourcify v2 + Blockscout v2, both keyless -- neither has verified source for this address)"


def build_review_guide_sheet(wb: Workbook, n_items: int) -> None:
    ws = wb.create_sheet("REVIEW_GUIDE", 0)
    ws.column_dimensions["A"].width = 115
    lines: list[tuple[str, str]] = [
        ("AuthGuard-7702 Pilot Code Review — Guide", "title"),
        ("", ""),
        ("What EIP-7702 authorization does", "h2"),
        ("EIP-7702 lets a normal wallet account (an EOA) temporarily run code from a "
         "separate 'delegate' contract, as if that code were the wallet's own. While "
         "delegated, that code can act using the wallet's own assets, token approvals, "
         "storage, and identity.", ""),
        ("", ""),
        ("The main question for every item", "h2"),
        (f"For each of the {n_items} Pilot contracts below: could an unauthorized person "
         "misuse this delegate to do something the wallet owner did not intend? Every row "
         "gives you the actual readable evidence needed to answer that -- a plain-English "
         "purpose, the specific sensitive functions found, a short readable code snippet or "
         "decompiled pseudocode showing exactly how each sensitive function is (or is not) "
         "restricted, and an AI-drafted preliminary answer.", ""),
        ("", ""),
        ("You do not need to read raw bytecode", "h2"),
        ("Every technical claim in this workbook is backed by a short, readable snippet in "
         "the relevant_code_snippet column -- either real decompiled pseudocode with a "
         "bytecode offset, or (when available) actual verified source. You are checking "
         "whether that snippet supports the AI's conclusion, not decoding raw bytecode "
         "yourself.", ""),
        ("", ""),
        ("How to choose SAFE, UNSAFE, or UNCERTAIN", "h2"),
        ("SAFE — the evidence shows sensitive actions are properly restricted (e.g. a clear "
         "'only the owner' or 'only the account itself' check on every sensitive function) "
         "and no concrete authorization-related vulnerability was identified.", ""),
        ("UNSAFE — the evidence identifies a concrete dangerous condition: for example, a "
         "sensitive function (asset transfer, arbitrary call, contract deployment, "
         "self-destruct) with NO caller restriction found anywhere in the contract, or an "
         "authorization check that uses tx.origin instead of msg.sender (a well-known "
         "phishing-vulnerable pattern, sometimes called 'SWC-115').", ""),
        ("UNCERTAIN — the evidence is not sufficient to make a reliable SAFE or UNSAFE "
         "decision (for example: an unresolved dependency, a contract too large/complex to "
         "fully trace, or a genuinely ambiguous finding). There is no penalty for choosing "
         "UNCERTAIN — it is often the correct, honest answer.", ""),
        ("", ""),
        ("Important warnings", "h2"),
        ("A contract merely being able to make external calls, or having a fallback "
         "function, does NOT by itself make it unsafe -- what matters is whether that "
         "capability is restricted to an authorized caller.", ""),
        ("An unresolved or unverified contract does NOT automatically mean unsafe -- but it "
         "does mean the finding should usually be UNCERTAIN unless other evidence (like a "
         "fully-traced access-control guard) is conclusive on its own.", ""),
        ("A contract using well-known libraries (OpenZeppelin Ownable/AccessControl, Safe) "
         "is a positive signal, but is not proof of safety by itself -- always check what "
         "the actual snippet shows for THIS deployment.", ""),
        ("The AI's proposed label and reasoning are a starting point, not the final answer "
         "-- they can be wrong. Your independent judgment matters.", ""),
        ("", ""),
        ("Workflow", "h2"),
        ("1. Read a row's contract_purpose, sensitive_function_names, and "
         "relevant_code_snippet.", ""),
        ("2. Read the AI's proposed label, confidence, and reasoning.", ""),
        ("3. Decide whether you agree; select SAFE, UNSAFE, or UNCERTAIN in your "
         "contributor_label column and write a short contributor_reason.", ""),
        ("4. Discuss difficult items with other contributors (group_discussion column).", ""),
        ("5. The lead author reads everything and records final_label / final_reason "
         "(those two columns are locked).", ""),
    ]
    row = 1
    for text, style in lines:
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if style == "title":
            cell.font = Font(bold=True, size=16)
        elif style == "h2":
            cell.font = Font(bold=True, size=12)
        row += 1
    ws.freeze_panes = "A2"


def build_review_items_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("REVIEW_ITEMS")

    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if col_name in EXTRA_WIDE_COLUMNS:
            width = 70
        elif col_name in WIDE_COLUMNS:
            width = 48
        else:
            width = 22
        if col_name in ("item_id", "code_source_or_decompiler_reference", "explorer_link"):
            width = 32
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    for row_idx, record in enumerate(rows, start=2):
        row_fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
        for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = LEAD_AUTHOR_FILL if col_name in LEAD_AUTHOR_COLUMNS else row_fill
            if col_name == "relevant_code_snippet":
                cell.font = CODE_FONT

    last_row = len(rows) + 1
    last_col = len(REVIEW_COLUMNS)

    def col_letter(name: str) -> str:
        return get_column_letter(REVIEW_COLUMNS.index(name) + 1)

    for col_name in LABEL_DROPDOWN_COLUMNS:
        dv = DataValidation(type="list", formula1=f'"{",".join(PRIMARY_LABELS)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter(col_name)}2:{col_letter(col_name)}{last_row}")

    ws.freeze_panes = ws.cell(row=2, column=2).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


def protect_lead_author_columns(ws, n_rows: int) -> None:
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        locked = col_name in LEAD_AUTHOR_COLUMNS
        for row in range(1, n_rows + 2):
            ws.cell(row=row, column=col_idx).protection = Protection(locked=locked)
    ws.protection = SheetProtection(
        sheet=True, password=None,
        formatCells=False, formatColumns=False, formatRows=False,
        insertRows=False, insertColumns=False, deleteRows=False, deleteColumns=False,
        sort=False, autoFilter=False, pivotTables=True,
        selectLockedCells=False, selectUnlockedCells=False,
    )


def main() -> int:
    manifest_rows = load_manifest_rows(MANIFEST_PATH)
    assert len(manifest_rows) == 20, f"expected 20 Pilot items, found {len(manifest_rows)}"

    packets_by_item = load_packets_by_item_id(EVIDENCE_DUMP_PATH, manifest_rows)

    with open(CODE_REVIEW_CONTENT_PATH) as f:
        content_by_item = json.load(f)["reviews"]

    rows = []
    for row in manifest_rows:
        item_id = row["item_id"]
        packet = packets_by_item[item_id]
        content = content_by_item[item_id]
        verification = load_verification_status(item_id)
        known = packet.get("known_project")

        impl_addr = None
        for cand_key in ("implementation_address",):
            pass
        # Pull implementation address out of the review content's proxy explanation if it
        # was resolved (recorded there in prose); fall back to explicit lookup fields.
        impl_resolved_text = content["proxy_and_upgrade_explanation"]

        record = {
            "item_id": item_id,
            "chain": row["chain"],
            "address": row["address"],
            "project_name": known["project"] if known else "(none documented)",
            "explorer_link": packet.get("explorer_link", ""),
            "verified_source_status": verified_source_status_text(verification),
            "implementation_address": _extract_impl_address(item_id),
            "implementation_resolved": impl_resolved_text,
            **{k: content[k] for k in [
                "contract_purpose", "sensitive_function_names", "relevant_code_snippet",
                "code_source_or_decompiler_reference", "access_control_explanation",
                "initialization_explanation", "asset_transfer_or_approval_explanation",
                "concrete_security_finding", "unresolved_questions",
                "ai_proposed_label", "ai_confidence", "ai_reasoning",
            ]},
        }
        record["proxy_and_upgrade_explanation"] = impl_resolved_text
        for c in CONTRIBUTOR_COLUMNS + GROUP_COLUMNS + LEAD_AUTHOR_COLUMNS:
            record.setdefault(c, "")
        rows.append(record)

    wb = Workbook()
    wb.remove(wb.active)
    build_review_guide_sheet(wb, n_items=len(manifest_rows))
    build_review_items_sheet(wb, rows)
    protect_lead_author_columns(wb["REVIEW_ITEMS"], n_rows=len(rows))

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"wrote {OUTPUT_PATH} ({len(rows)} items)")
    return 0


def _extract_impl_address(item_id: str) -> str:
    folder = os.path.join(CODE_EVIDENCE_DIR, item_id.replace(":", "_"))
    readme_path = os.path.join(folder, "README.md")
    if not os.path.exists(readme_path):
        return "N/A (not a proxy)"
    with open(readme_path) as f:
        text = f.read()
    marker = "implementation address "
    if marker in text:
        after = text.split(marker, 1)[1]
        addr = after.split()[0].rstrip(".")
        return addr
    return "N/A (not a proxy)"


if __name__ == "__main__":
    sys.exit(main())
