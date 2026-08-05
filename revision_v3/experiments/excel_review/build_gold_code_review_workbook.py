"""Builds Gold_Dev_Code_Review.xlsx / Gold_Test_Code_Review.xlsx -- same simplified 2-sheet
structure as Pilot_Code_Review.xlsx, populated from the automated provisional-label records
(revision_v3/results/llm_provisional/{set}_labels.json) plus a real, offset-anchored
disassembly snippet pulled from each item's saved decompiled/disassembly.txt.

Does not modify Pilot_Code_Review.xlsx or any Phase 3A file.

Usage:
    python3 revision_v3/experiments/excel_review/build_gold_code_review_workbook.py --sample-set gold_dev
    python3 revision_v3/experiments/excel_review/build_gold_code_review_workbook.py --sample-set gold_test
"""
from __future__ import annotations

import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
HUMAN_EVAL_DIR = os.path.join(REPO_ROOT, "revision_v3", "human_eval")
RESULTS_DIR = os.path.join(REPO_ROOT, "revision_v3", "results", "llm_provisional")
sys.path.insert(0, HUMAN_EVAL_DIR)
from taxonomy import PRIMARY_LABELS  # noqa: E402

EVIDENCE_DIRS = {"gold_dev": "gold_dev_code_evidence", "gold_test": "gold_test_code_evidence"}
OUTPUT_NAMES = {"gold_dev": "Gold_Dev_Code_Review.xlsx", "gold_test": "Gold_Test_Code_Review.xlsx"}
TITLE_NAMES = {"gold_dev": "Gold-Dev", "gold_test": "Gold-Test"}

EVIDENCE_COLUMNS = [
    "item_id", "chain", "address", "explorer_link", "verified_source_status",
    "contract_purpose", "sensitive_functions", "relevant_code_snippet",
    "code_source_or_decompiler_reference", "access_control_analysis",
    "initialization_analysis", "proxy_and_upgrade_analysis", "asset_operation_analysis",
    "authorization_specific_analysis", "concrete_finding", "unresolved_questions",
]
AI_COLUMNS = ["llm_provisional_label", "llm_provisional_confidence", "llm_provisional_reason_category", "ai_reasoning"]
CONTRIBUTOR_COLUMNS = ["contributor_label", "contributor_rationale"]
GROUP_COLUMNS = ["group_discussion"]
LEAD_AUTHOR_COLUMNS = ["final_label", "final_rationale"]
REVIEW_COLUMNS = EVIDENCE_COLUMNS + AI_COLUMNS + CONTRIBUTOR_COLUMNS + GROUP_COLUMNS + LEAD_AUTHOR_COLUMNS
LABEL_DROPDOWN_COLUMNS = ["contributor_label", "final_label"]
WIDE_COLUMNS = {
    "contract_purpose", "sensitive_functions", "code_source_or_decompiler_reference",
    "access_control_analysis", "initialization_analysis", "proxy_and_upgrade_analysis",
    "asset_operation_analysis", "authorization_specific_analysis", "concrete_finding",
    "unresolved_questions", "ai_reasoning", "contributor_rationale", "group_discussion",
    "final_rationale",
}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ROW_FILL_EVEN = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
LEAD_AUTHOR_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CODE_FONT = Font(name="Consolas", size=9)

EXPLORER_HOSTS = {
    "ethereum": "etherscan.io", "optimism": "optimistic.etherscan.io", "base": "basescan.org",
    "arbitrum": "arbiscan.io", "polygon": "polygonscan.com", "bnb": "bscscan.com",
    "gnosis": "gnosisscan.io",
}


def extract_snippet(evidence_dir: str, item_id: str, per_function: list) -> str:
    folder = os.path.join(evidence_dir, item_id.replace(":", "_"), "decompiled")
    disasm_path = os.path.join(folder, "disassembly.txt")
    if not os.path.exists(disasm_path) or not per_function:
        return "(no dispatched function found to anchor a snippet -- see full disassembly.txt)"
    with open(disasm_path) as f:
        lines = f.read().splitlines()
    offsets = [int(l[:5]) for l in lines]

    # Prefer the offset of a GUARDED or OPEN sensitive (state-changing) function; fall back to
    # the first function's dispatch offset.
    anchor_fn = next((fn for fn in per_function if fn.get("state_mutability") in ("nonpayable", "payable")
                       and fn.get("guard_status") in ("GUARDED", "OPEN")), per_function[0])
    anchor_offset = anchor_fn.get("guard_offset") or anchor_fn["bytecode_offset"]
    try:
        idx = offsets.index(anchor_offset)
    except ValueError:
        idx = next((i for i, o in enumerate(offsets) if o >= anchor_offset), 0)
    lo, hi = max(0, idx - 8), min(len(lines), idx + 12)
    header = (f"-- {anchor_fn.get('resolved_signature') or anchor_fn['selector']} "
              f"(guard_status={anchor_fn.get('guard_status')}) --")
    return header + "\n" + "\n".join(lines[lo:hi])


def build_review_guide_sheet(wb: Workbook, sample_set: str, n_items: int) -> None:
    ws = wb.create_sheet("REVIEW_GUIDE", 0)
    ws.column_dimensions["A"].width = 115
    title = TITLE_NAMES[sample_set]
    lines: list[tuple[str, str]] = [
        (f"AuthGuard-7702 {title} Code Review — Guide", "title"),
        ("", ""),
        ("What this workbook is", "h2"),
        (f"This workbook covers all {n_items} frozen {title} items. Every row's label in the "
         "llm_provisional_label column is a PROVISIONAL, LLM-generated reference label -- "
         "not a human label, not ground truth. It exists so the technical research pipeline "
         "can proceed while independent human review happens on its own timeline. Your "
         "review here (contributor_label / final_label) is what actually matters.", ""),
        ("", ""),
        ("What EIP-7702 authorization does", "h2"),
        ("EIP-7702 lets a normal wallet account (an EOA) temporarily run code from a "
         "separate 'delegate' contract, as if that code were the wallet's own. The main "
         "question for every item: could an unauthorized person misuse this delegate?", ""),
        ("", ""),
        ("You do not need to read raw bytecode", "h2"),
        ("Every row's relevant_code_snippet column contains a short, real, offset-anchored "
         "excerpt of the decompiled disassembly around the specific function the finding is "
         "about (or, when verified source exists, real source). You are checking whether that "
         "snippet supports the stated finding, not decoding bytecode yourself.", ""),
        ("", ""),
        ("How to choose SAFE, UNSAFE, or UNCERTAIN", "h2"),
        ("SAFE — sensitive actions are properly restricted; no concrete authorization-related "
         "danger was identified.", ""),
        ("UNSAFE — a concrete dangerous condition was identified: e.g. an asset-moving or "
         "arbitrary-call function with no caller restriction found anywhere in its traced "
         "body, or a tx.origin-based authorization check guarding a high-impact action.", ""),
        ("UNCERTAIN — the evidence is insufficient, ambiguous, or unresolved (e.g. an "
         "unresolved proxy implementation, or bytecode too atypical to reliably trace). "
         "There is no penalty for UNCERTAIN.", ""),
        ("", ""),
        ("Important warnings", "h2"),
        ("A capability existing (CALL, DELEGATECALL, fallback, token selectors) is not "
         "itself unsafe -- what matters is whether it is restricted to an authorized caller.", ""),
        ("The provisional label was produced by an automated guard-tracer plus templated "
         "LLM write-up, not a full line-by-line manual audit like the original 20-item Pilot "
         "batch -- treat it as a well-evidenced starting point, not a final answer.", ""),
        ("An unverified/unresolved contract is not automatically unsafe, but usually warrants "
         "UNCERTAIN unless another finding is independently conclusive.", ""),
        ("", ""),
        ("Workflow", "h2"),
        ("1. Read contract_purpose, sensitive_functions, and relevant_code_snippet.", ""),
        ("2. Read the LLM's provisional label, confidence, and reasoning.", ""),
        ("3. Select SAFE, UNSAFE, or UNCERTAIN in contributor_label; write a short "
         "contributor_rationale.", ""),
        ("4. Discuss difficult items with other contributors (group_discussion).", ""),
        ("5. The lead author records final_label / final_rationale (locked columns).", ""),
    ]
    row = 1
    for text, style in lines:
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if style == "title":
            cell.font = Font(bold=True, size=16)
        elif style == "h2":
            cell.font = Font(bold=True, size=12)
        row += 1
    ws.freeze_panes = "A2"


def build_review_items_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("REVIEW_ITEMS")
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if col_name == "relevant_code_snippet":
            width = 70
        elif col_name in WIDE_COLUMNS:
            width = 45
        else:
            width = 22
        if col_name in ("item_id", "explorer_link", "code_source_or_decompiler_reference"):
            width = 30
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    for row_idx, record in enumerate(rows, start=2):
        row_fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
        for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = LEAD_AUTHOR_FILL if col_name in LEAD_AUTHOR_COLUMNS else row_fill
            if col_name == "relevant_code_snippet":
                cell.font = CODE_FONT

    last_row = len(rows) + 1

    def col_letter(name):
        return get_column_letter(REVIEW_COLUMNS.index(name) + 1)

    for col_name in LABEL_DROPDOWN_COLUMNS:
        dv = DataValidation(type="list", formula1=f'"{",".join(PRIMARY_LABELS)}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter(col_name)}2:{col_letter(col_name)}{last_row}")

    ws.freeze_panes = ws.cell(row=2, column=2).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REVIEW_COLUMNS))}{last_row}"


def protect_lead_author_columns(ws, n_rows: int) -> None:
    for col_idx, col_name in enumerate(REVIEW_COLUMNS, start=1):
        locked = col_name in LEAD_AUTHOR_COLUMNS
        for row in range(1, n_rows + 2):
            ws.cell(row=row, column=col_idx).protection = Protection(locked=locked)
    ws.protection = SheetProtection(
        sheet=True, password=None, formatCells=False, formatColumns=False, formatRows=False,
        insertRows=False, insertColumns=False, deleteRows=False, deleteColumns=False,
        sort=False, autoFilter=False, pivotTables=True,
        selectLockedCells=False, selectUnlockedCells=False,
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--sample-set", required=True, choices=list(EVIDENCE_DIRS))
    args = parser.parse_args()

    evidence_dir = os.path.join(HUMAN_EVAL_DIR, EVIDENCE_DIRS[args.sample_set])
    with open(os.path.join(evidence_dir, "evidence_manifest.json")) as f:
        evidence_manifest = json.load(f)["items"]
    with open(os.path.join(RESULTS_DIR, f"{args.sample_set}_labels.json")) as f:
        label_records = {r["item_id"]: r for r in json.load(f)["records"]}

    rows = []
    for item_id, rec in label_records.items():
        ev = evidence_manifest[item_id]
        per_fn = ev["guard_trace_summary"]["per_function"]
        verification = ev["verification"]
        chain, address = rec["chain"], rec["address"]
        ai_reasoning = (
            f"{rec['access_control_analysis']} {rec['authorization_specific_analysis']} "
            f"Confidence rationale: label={rec['llm_provisional_label']} "
            f"({rec['llm_provisional_confidence']}), alternative={rec['alternative_plausible_label']} "
            f"{rec['alternative_label_condition']}."
        )
        rows.append({
            "item_id": item_id, "chain": chain, "address": address,
            "explorer_link": f"https://{EXPLORER_HOSTS.get(chain, '')}/address/{address}",
            "verified_source_status": (
                "VERIFIED" if verification["verified"] else "NOT VERIFIED (Sourcify v2 + Blockscout v2 checked live)"
            ),
            "contract_purpose": rec["contract_purpose"],
            "sensitive_functions": rec["sensitive_functions"],
            "relevant_code_snippet": extract_snippet(evidence_dir, item_id, per_fn),
            "code_source_or_decompiler_reference": rec["evidence_references"],
            "access_control_analysis": rec["access_control_analysis"],
            "initialization_analysis": rec["initialization_analysis"],
            "proxy_and_upgrade_analysis": rec["proxy_and_upgrade_analysis"],
            "asset_operation_analysis": rec["asset_operation_analysis"],
            "authorization_specific_analysis": rec["authorization_specific_analysis"],
            "concrete_finding": rec["concrete_finding"],
            "unresolved_questions": rec["unresolved_questions"],
            "llm_provisional_label": rec["llm_provisional_label"],
            "llm_provisional_confidence": rec["llm_provisional_confidence"],
            "llm_provisional_reason_category": rec["llm_provisional_reason_category"],
            "ai_reasoning": ai_reasoning,
            "contributor_label": "", "contributor_rationale": "",
            "group_discussion": "", "final_label": "", "final_rationale": "",
        })

    wb = Workbook()
    wb.remove(wb.active)
    build_review_guide_sheet(wb, args.sample_set, n_items=len(rows))
    build_review_items_sheet(wb, rows)
    protect_lead_author_columns(wb["REVIEW_ITEMS"], n_rows=len(rows))

    out_path = os.path.join(HUMAN_EVAL_DIR, OUTPUT_NAMES[args.sample_set])
    wb.save(out_path)
    print(f"wrote {out_path} ({len(rows)} items)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
