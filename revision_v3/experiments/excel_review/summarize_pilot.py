"""Phase 3A, Part 10: summarizes completed contributor reviews and final lead-author labels
from Pilot_Master_Adjudication.xlsx. Writes revision_v3/reports/PILOT_REVIEW_SUMMARY.md.

Does NOT calculate ML accuracy -- this is explicitly out of scope for the Pilot (per the audit
brief). If zero contributor sections have been imported, the report is written with status
PENDING_REVIEWS rather than fabricating statistics from nothing.
"""
from __future__ import annotations

import os
import sys
from collections import Counter

import openpyxl

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
MASTER_PATH = os.path.join(REPO_ROOT, "revision_v3", "human_eval", "Pilot_Master_Adjudication.xlsx")
OUT_PATH = os.path.join(REPO_ROOT, "revision_v3", "reports", "PILOT_REVIEW_SUMMARY.md")


def load_master():
    wb = openpyxl.load_workbook(MASTER_PATH)
    ws = wb["MASTER_ITEMS"]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({h: ws.cell(row=r, column=i + 1).value for i, h in enumerate(headers)})
    NON_CONTRIBUTOR_LABEL_COLUMNS = {"llm_proposed_label", "final_label"}
    contributor_prefixes = sorted({
        h.rsplit("_", 1)[0] for h in headers
        if h and h.endswith("_label") and h not in NON_CONTRIBUTOR_LABEL_COLUMNS
    })
    log_rows = []
    if "_IMPORT_LOG" in wb.sheetnames:
        log = wb["_IMPORT_LOG"]
        for r in range(2, log.max_row + 1):
            if log.cell(row=r, column=1).value:
                log_rows.append([c.value for c in log[r]])
    return rows, contributor_prefixes, log_rows


def write_pending_report(n_items: int) -> None:
    content = f"""# Pilot Review Summary

**Status: PENDING_REVIEWS**

`Pilot_Master_Adjudication.xlsx` has {n_items} items seeded with evidence and LLM preliminary
analysis, but zero contributor responses have been imported yet (`_IMPORT_LOG` is empty). This
report will be regenerated with real statistics once
`revision_v3/human_eval/import_reviewer_workbook.py` has been run for at least one completed
contributor file.

No numbers are fabricated or estimated in this placeholder -- re-run
`revision_v3/experiments/excel_review/summarize_pilot.py` after imports to populate this
report.
"""
    with open(OUT_PATH, "w") as f:
        f.write(content)
    print(f"wrote {OUT_PATH} (PENDING_REVIEWS)")


def main() -> int:
    rows, contributor_prefixes, log_rows = load_master()
    n_items = len(rows)

    if not contributor_prefixes or not log_rows:
        write_pending_report(n_items)
        return 0

    lines = ["# Pilot Review Summary", ""]
    lines.append(f"**Status: IN_PROGRESS** ({len(contributor_prefixes)} contributor(s) imported)")
    lines.append("")

    lines.append("## Contributor completion counts")
    for prefix in contributor_prefixes:
        n_completed = sum(1 for r in rows if r.get(f"{prefix}_label"))
        lines.append(f"- {prefix}: {n_completed} / {n_items} items labeled")
    lines.append("")

    lines.append("## Initial label distribution (per contributor)")
    for prefix in contributor_prefixes:
        counts = Counter(r.get(f"{prefix}_label") for r in rows if r.get(f"{prefix}_label"))
        lines.append(f"- {prefix}: {dict(counts)}")
    lines.append("")

    lines.append("## Initial agreement rate (unanimous vs. disagreement, across imported contributors)")
    unanimous = sum(1 for r in rows if str(r.get("disagreement_summary", "")).startswith("unanimous"))
    disagree = sum(1 for r in rows if str(r.get("disagreement_summary", "")).startswith("DISAGREEMENT"))
    lines.append(f"- Unanimous: {unanimous} / {n_items}")
    lines.append(f"- Disagreement: {disagree} / {n_items}")
    lines.append("")

    lines.append("## Contributor agreement with LLM")
    agree_col_candidates = [c for c in rows[0].keys() if c == "agree_with_llm"]
    # agree_with_llm lives only in the per-reviewer ORIGINAL file, not persisted per-contributor
    # in the master by column name (only label/reason/confidence/rationale sections are merged).
    lines.append("- Not computed from the master workbook (agree_with_llm is recorded per "
                 "contributor file but not currently merged into MASTER_ITEMS as a named "
                 "per-contributor column; re-run the importer with an extended column set if "
                 "this statistic is needed before the Pilot concludes).")
    lines.append("")

    lines.append("## Common disagreement reason categories")
    disagreement_reasons = Counter()
    for r in rows:
        if str(r.get("disagreement_summary", "")).startswith("DISAGREEMENT"):
            for prefix in contributor_prefixes:
                reason = r.get(f"{prefix}_reason_category")
                if reason:
                    disagreement_reasons[reason] += 1
    lines.append(f"- {dict(disagreement_reasons.most_common())}" if disagreement_reasons else "- (none yet)")
    lines.append("")

    lines.append("## Final lead-author labels")
    final_counts = Counter(r.get("final_label") for r in rows if r.get("final_label"))
    n_finalized = sum(final_counts.values())
    lines.append(f"- Finalized: {n_finalized} / {n_items}")
    lines.append(f"- SAFE: {final_counts.get('SAFE', 0)}")
    lines.append(f"- UNSAFE: {final_counts.get('UNSAFE', 0)}")
    lines.append(f"- UNCERTAIN: {final_counts.get('UNCERTAIN', 0)}")
    lines.append("")

    lines.append("## Average contributor confidence")
    conf_map = {"high": 3, "medium": 2, "low": 1}
    conf_values = []
    for r in rows:
        for prefix in contributor_prefixes:
            c = r.get(f"{prefix}_confidence")
            if c in conf_map:
                conf_values.append(conf_map[c])
    if conf_values:
        lines.append(f"- Mean confidence (low=1,medium=2,high=3): {sum(conf_values)/len(conf_values):.2f} "
                     f"(n={len(conf_values)})")
    else:
        lines.append("- (no confidence values recorded yet)")
    lines.append("")

    lines.append("## Cases where the final decision disagrees with the LLM")
    disagree_with_llm = [r["item_id"] for r in rows
                         if r.get("final_label") and r.get("llm_proposed_label")
                         and r["final_label"] != r["llm_proposed_label"]]
    lines.append(f"- {len(disagree_with_llm)} item(s): {disagree_with_llm}" if disagree_with_llm else "- none yet finalized to compare")
    lines.append("")

    lines.append("## Cases requiring additional evidence (flagged by any contributor or the LLM)")
    flagged = [r["item_id"] for r in rows if r.get("llm_points_to_verify")]
    lines.append(f"- {len(flagged)} item(s) have LLM-flagged verification points (see llm_points_to_verify "
                 "column in MASTER_ITEMS for detail)")
    lines.append("")

    lines.append("## Import log")
    for row in log_rows:
        lines.append(f"- {row[0]} imported {row[3]} items from `{row[1]}` at {row[2]}")

    with open(OUT_PATH, "w") as f:
        f.write("\n".join(lines) + "\n")
    print(f"wrote {OUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
