"""Shared Excel workbook construction helpers for Phase 3A (Pilot) and, later, Gold-Dev /
Gold-Test. One module so every workbook (Pilot, Gold-Dev, Gold-Test, and any reviewer copy or
master-adjudication derivative) has identical structure, taxonomy, and validation rules.
"""
from __future__ import annotations

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
sys.path.insert(0, os.path.join(REPO_ROOT, "revision_v3", "human_eval"))
sys.path.insert(0, os.path.join(REPO_ROOT, "revision_v3", "src"))

from taxonomy import (  # noqa: E402
    AGREEMENT_VALUES, ALL_REASON_CATEGORIES, CONFIDENCE_LEVELS,
    INCLUDED_IN_BINARY_EVALUATION_VALUES, PRIMARY_LABELS, REASONS_BY_LABEL,
    SAFE_REASONS, UNCERTAIN_REASONS, UNSAFE_REASONS,
)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
EVIDENCE_FILL = PatternFill(start_color="EAF1F8", end_color="EAF1F8", fill_type="solid")
LLM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CONTRIBUTOR_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LEAD_AUTHOR_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
LOCK_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

EVIDENCE_COLUMNS = [
    "item_id", "chain", "contract_address", "project_name_if_known", "documentation_url",
    "explorer_url", "verified_source_status", "runtime_bytecode_hash", "runtime_bytecode_size",
    "proxy_or_designator_status", "resolved_implementation",
    "ownership_and_initialization_summary", "external_call_summary",
    "asset_transfer_and_approval_summary", "delegatecall_and_upgrade_summary",
    "authorization_history_summary", "other_security_evidence",
    "evidence_packet_path_or_link",
]

LLM_COLUMNS = [
    "llm_proposed_label", "llm_reason_category", "llm_confidence", "llm_rationale",
    "llm_evidence", "llm_uncertainty", "llm_points_to_verify",
]

CONTRIBUTOR_COLUMNS = [
    "contributor_name", "contributor_label", "contributor_reason_category",
    "contributor_confidence", "agree_with_llm", "contributor_rationale",
    "important_evidence", "questions_or_discussion_notes", "review_date",
]

LEAD_AUTHOR_COLUMNS = [
    "final_label", "final_reason_category", "final_rationale", "final_decision_date",
    "included_in_binary_evaluation",
]

ALL_ITEM_COLUMNS = EVIDENCE_COLUMNS + LLM_COLUMNS + CONTRIBUTOR_COLUMNS + LEAD_AUTHOR_COLUMNS

COLUMN_GROUP_FILL = {}
for c in EVIDENCE_COLUMNS:
    COLUMN_GROUP_FILL[c] = EVIDENCE_FILL
for c in LLM_COLUMNS:
    COLUMN_GROUP_FILL[c] = LLM_FILL
for c in CONTRIBUTOR_COLUMNS:
    COLUMN_GROUP_FILL[c] = CONTRIBUTOR_FILL
for c in LEAD_AUTHOR_COLUMNS:
    COLUMN_GROUP_FILL[c] = LEAD_AUTHOR_FILL


def _write_markdown_as_sheet(ws, title: str, markdown_path: str) -> None:
    ws.column_dimensions["A"].width = 130
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    row = 3
    with open(markdown_path) as f:
        for line in f:
            line = line.rstrip("\n")
            cell = ws.cell(row=row, column=1, value=line if line else " ")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if line.startswith("# "):
                cell.font = Font(bold=True, size=14)
            elif line.startswith("## "):
                cell.font = Font(bold=True, size=12)
            elif line.startswith("### "):
                cell.font = Font(bold=True, size=11)
            row += 1
    ws.freeze_panes = "A2"


def build_start_here_sheet(wb: Workbook, review_set_name: str, n_items: int) -> None:
    ws = wb.create_sheet("START_HERE", 0)
    ws.column_dimensions["A"].width = 110
    lines = [
        (f"AuthGuard-7702 {review_set_name} Review — START HERE", "title"),
        ("", ""),
        ("Purpose", "h2"),
        (f"This workbook asks you to review {n_items} EIP-7702 delegate contracts and decide, "
         "for each, whether the available security evidence supports SAFE, UNSAFE, or "
         "UNCERTAIN. Your independent judgment — not the AI assistant's preliminary "
         "analysis — is what this review needs.", ""),
        ("", ""),
        ("Simple workflow", "h2"),
        ("1. Read the EIP7702_GUIDE sheet once (5-10 minutes) if you're new to EIP-7702.", ""),
        ("2. Read the REVIEW_CHECKLIST sheet and keep it open while you work.", ""),
        ("3. Skim the EXAMPLES sheet to calibrate what SAFE / UNSAFE / UNCERTAIN look like.", ""),
        ("4. Go to the PILOT_ITEMS sheet (or GOLD_DEV_ITEMS / GOLD_TEST_ITEMS). For each row:", ""),
        ("     a. Read the evidence columns (blue).", ""),
        ("     b. Read the LLM's preliminary analysis (yellow) — treat it as a colleague's "
         "first draft, not an answer key.", ""),
        ("     c. Fill in YOUR OWN judgment in the green contributor columns.", ""),
        ("5. Save the file when done (see below) and return it as instructed.", ""),
        ("", ""),
        ("Label definitions", "h2"),
        ("SAFE — the available evidence supports that the delegate is appropriate to "
         "authorize under EIP-7702; no concrete authorization-related security risk was "
         "identified.", ""),
        ("UNSAFE — the available evidence shows a concrete security risk that could make "
         "authorization dangerous.", ""),
        ("UNCERTAIN — the available evidence is insufficient to make a reliable SAFE or "
         "UNSAFE decision (including cases that simply can't be assessed from the runtime "
         "bytecode available). There is no penalty for choosing UNCERTAIN.", ""),
        ("", ""),
        ("Saving your file", "h2"),
        ("Use File > Save (keep the .xlsx format). Do not rename the file. When done, return "
         "it to the lead author exactly as instructed in the email/message you received.", ""),
        ("", ""),
        ("A reminder", "h2"),
        ("The lead author reads every contributor's response and every discussion note, then "
         "makes the FINAL decision themselves (the orange columns, at the far right of the "
         "items sheet). Those columns are locked — you cannot and should not edit them. Your "
         "job is to give your best independent judgment, not to guess what the final answer "
         "will be.", ""),
    ]
    row = 1
    for text, style in lines:
        cell = ws.cell(row=row, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if style == "title":
            cell.font = Font(bold=True, size=16)
        elif style == "h2":
            cell.font = Font(bold=True, size=12)
        row += 1
    ws.freeze_panes = "A2"


def build_guide_sheet(wb: Workbook) -> None:
    guide_path = os.path.join(REPO_ROOT, "revision_v3", "human_eval", "REVIEWER_GUIDE.md")
    ws = wb.create_sheet("EIP7702_GUIDE")
    _write_markdown_as_sheet(ws, "EIP-7702 Reviewer Guide", guide_path)


def build_checklist_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("REVIEW_CHECKLIST")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 120
    ws["A1"] = "Step"
    ws["B1"] = "Review Checklist (work through in order)"
    for c in ("A1", "B1"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL

    guide_path = os.path.join(REPO_ROOT, "revision_v3", "human_eval", "REVIEWER_GUIDE.md")
    with open(guide_path) as f:
        content = f.read()
    section = content.split("## 3. Review checklist")[1].split("## 4. Educational examples")[0]

    row = 2
    for line in section.split("\n"):
        line = line.rstrip()
        if not line.strip():
            continue
        ws.cell(row=row, column=1, value="")
        cell = ws.cell(row=row, column=2, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("### "):
            cell.font = Font(bold=True, size=12)
        elif line.startswith("**Important"):
            cell.font = Font(bold=True, color="C00000")
        row += 1
    ws.freeze_panes = "B2"


def build_examples_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("EXAMPLES")
    ws.column_dimensions["A"].width = 130
    guide_path = os.path.join(REPO_ROOT, "revision_v3", "human_eval", "REVIEWER_GUIDE.md")
    with open(guide_path) as f:
        content = f.read()
    section = content.split("## 4. Educational examples")[1].split("## 5. A note")[0]
    ws["A1"] = "Educational Examples (synthetic — not real Pilot/Gold-Dev/Gold-Test items)"
    ws["A1"].font = Font(bold=True, size=14)
    row = 3
    for line in section.split("\n"):
        cell = ws.cell(row=row, column=1, value=line if line.strip() else " ")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("### "):
            cell.font = Font(bold=True, size=12)
        elif line.startswith("**Correct label**") or line.startswith("**Common incorrect"):
            cell.font = Font(bold=True)
        row += 1
    ws.freeze_panes = "A2"


def _evidence_row_from_manifest_row(row: dict, packet: dict) -> dict:
    proxy = packet.get("proxy_evidence", {})
    struct = packet.get("structural", {})
    storage = packet.get("storage_operations", {})
    token = packet.get("token_transfer_evidence", {})
    known_project = packet.get("known_project")

    ownership_bits = []
    if proxy.get("any_admin_ownership_selector_present"):
        present = [k for k, v in proxy.get("admin_ownership_selectors_present", {}).items() if v]
        ownership_bits.append(f"admin/ownership selectors detected: {', '.join(present)}")
    else:
        ownership_bits.append("no named admin/ownership selector detected")
    ownership_bits.append(f"SSTORE={storage.get('n_sstore')}, SLOAD={storage.get('n_sload')}")
    if proxy.get("is_eip7702_designator"):
        ownership_bits.append(f"EIP-7702 designator -> {proxy.get('designator_target_address')}")

    external_call_bits = [
        f"CALL={struct.get('n_call')}, STATICCALL={struct.get('n_staticcall')}, "
        f"DELEGATECALL={struct.get('n_delegatecall')}, CALLCODE={struct.get('n_callcode')}",
    ]

    asset_bits = []
    moved = [k for k, v in token.get("token_movement_selectors_present", {}).items() if v]
    if moved:
        asset_bits.append(f"token-movement selectors: {', '.join(moved)}")
    if token.get("approval_selector_present"):
        asset_bits.append("approve(address,uint256) selector present")
    if not asset_bits:
        asset_bits.append("no token-transfer/approval selector detected")

    delegatecall_bits = []
    if proxy.get("has_delegatecall"):
        delegatecall_bits.append(
            f"{proxy.get('delegatecall_count')} DELEGATECALL instruction(s), last at "
            f"position ratio {proxy.get('delegatecall_last_position_ratio')}"
        )
        if proxy.get("resembles_minimal_forwarder"):
            delegatecall_bits.append("resembles a minimal forwarder pattern")
        for slot_name in ("eip1967_implementation_slot_present", "eip1967_admin_slot_present",
                          "eip1967_beacon_slot_present"):
            if proxy.get(slot_name):
                delegatecall_bits.append(f"{slot_name}=True")
    else:
        delegatecall_bits.append("no DELEGATECALL detected")

    other_bits = []
    if struct.get("n_selfdestruct", 0) > 0:
        other_bits.append(f"SELFDESTRUCT count={struct.get('n_selfdestruct')}")
    if struct.get("has_sensitive_selector"):
        other_bits.append(f"{struct.get('n_sensitive_selectors')} sensitive-name-list selector(s) present")
    other_bits.append(packet.get("deterministic_summary", ""))

    return {
        "item_id": row["item_id"],
        "chain": row["chain"],
        "contract_address": row["address"],
        "project_name_if_known": known_project["project"] if known_project else "(none documented)",
        "documentation_url": known_project["documentation_url"] if known_project else "",
        "explorer_url": packet.get("explorer_link", ""),
        "verified_source_status": packet.get("verified_source_code_availability", {}).get("status", ""),
        "runtime_bytecode_hash": packet.get("runtime_bytecode_sha256", ""),
        "runtime_bytecode_size": packet.get("runtime_bytecode_length_bytes", ""),
        "proxy_or_designator_status": (
            "EIP-7702 designator" if proxy.get("is_eip7702_designator")
            else ("DELEGATECALL-based proxy pattern" if proxy.get("has_delegatecall") else "not a designator/proxy pattern")
        ),
        "resolved_implementation": proxy.get("designator_target_address") or "(not resolvable offline)",
        "ownership_and_initialization_summary": "; ".join(ownership_bits),
        "external_call_summary": "; ".join(external_call_bits),
        "asset_transfer_and_approval_summary": "; ".join(asset_bits),
        "delegatecall_and_upgrade_summary": "; ".join(delegatecall_bits),
        "authorization_history_summary": packet.get("authorization_history", {}).get("status", ""),
        "other_security_evidence": "; ".join(other_bits),
        "evidence_packet_path_or_link": "revision_v3/human_eval/llm_reviews/pilot_evidence_dump.json",
    }


def build_items_sheet(wb: Workbook, sheet_name: str, manifest_rows: list, packets_by_item: dict,
                       llm_reviews: dict) -> None:
    ws = wb.create_sheet(sheet_name)

    for col_idx, col_name in enumerate(ALL_ITEM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        width = 22
        if col_name in ("ownership_and_initialization_summary", "external_call_summary",
                        "asset_transfer_and_approval_summary", "delegatecall_and_upgrade_summary",
                        "other_security_evidence", "llm_rationale", "llm_evidence",
                        "llm_uncertainty", "llm_points_to_verify", "contributor_rationale",
                        "important_evidence", "questions_or_discussion_notes", "final_rationale"):
            width = 45
        elif col_name in ("item_id", "runtime_bytecode_hash", "documentation_url", "explorer_url",
                          "evidence_packet_path_or_link"):
            width = 30
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(manifest_rows, start=2):
        packet = packets_by_item[row["item_id"]]
        evidence = _evidence_row_from_manifest_row(row, packet)
        llm = llm_reviews.get(row["item_id"], {})

        record = dict(evidence)
        record["llm_proposed_label"] = llm.get("llm_proposed_label", "")
        record["llm_reason_category"] = llm.get("llm_reason_category", "")
        record["llm_confidence"] = llm.get("llm_confidence", "")
        record["llm_rationale"] = llm.get("llm_rationale", "")
        record["llm_evidence"] = llm.get("llm_evidence", "")
        record["llm_uncertainty"] = llm.get("llm_uncertainty", "")
        record["llm_points_to_verify"] = llm.get("llm_points_to_verify", "")
        for c in CONTRIBUTOR_COLUMNS + LEAD_AUTHOR_COLUMNS:
            record.setdefault(c, "")

        for col_idx, col_name in enumerate(ALL_ITEM_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = COLUMN_GROUP_FILL.get(col_name, PatternFill())

    last_row = len(manifest_rows) + 1
    last_col = len(ALL_ITEM_COLUMNS)

    def col_letter(name):
        return get_column_letter(ALL_ITEM_COLUMNS.index(name) + 1)

    dv_label = DataValidation(type="list", formula1=f'"{",".join(PRIMARY_LABELS)}"', allow_blank=True)
    ws.add_data_validation(dv_label)
    dv_label.add(f"{col_letter('contributor_label')}2:{col_letter('contributor_label')}{last_row}")

    dv_reason = DataValidation(type="list", formula1=f'"{",".join(ALL_REASON_CATEGORIES)}"', allow_blank=True)
    ws.add_data_validation(dv_reason)
    dv_reason.add(f"{col_letter('contributor_reason_category')}2:{col_letter('contributor_reason_category')}{last_row}")

    dv_conf = DataValidation(type="list", formula1=f'"{",".join(CONFIDENCE_LEVELS)}"', allow_blank=True)
    ws.add_data_validation(dv_conf)
    dv_conf.add(f"{col_letter('contributor_confidence')}2:{col_letter('contributor_confidence')}{last_row}")

    dv_agree = DataValidation(type="list", formula1=f'"{",".join(AGREEMENT_VALUES)}"', allow_blank=True)
    ws.add_data_validation(dv_agree)
    dv_agree.add(f"{col_letter('agree_with_llm')}2:{col_letter('agree_with_llm')}{last_row}")

    dv_final_label = DataValidation(type="list", formula1=f'"{",".join(PRIMARY_LABELS)}"', allow_blank=True)
    ws.add_data_validation(dv_final_label)
    dv_final_label.add(f"{col_letter('final_label')}2:{col_letter('final_label')}{last_row}")

    dv_final_reason = DataValidation(type="list", formula1=f'"{",".join(ALL_REASON_CATEGORIES)}"', allow_blank=True)
    ws.add_data_validation(dv_final_reason)
    dv_final_reason.add(f"{col_letter('final_reason_category')}2:{col_letter('final_reason_category')}{last_row}")

    dv_included = DataValidation(type="list", formula1=f'"{",".join(INCLUDED_IN_BINARY_EVALUATION_VALUES)}"', allow_blank=True)
    ws.add_data_validation(dv_included)
    dv_included.add(f"{col_letter('included_in_binary_evaluation')}2:{col_letter('included_in_binary_evaluation')}{last_row}")

    ws.freeze_panes = ws.cell(row=2, column=len(EVIDENCE_COLUMNS) + 1).coordinate
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.row_dimensions[1].height = 30


def protect_lead_author_columns(ws, n_rows: int, sheet_password: str | None = None) -> None:
    """Locks the final_* columns so contributors can't accidentally edit them; everything
    else stays editable. Requires enabling worksheet protection."""
    from openpyxl.worksheet.protection import SheetProtection

    for col_idx, col_name in enumerate(ALL_ITEM_COLUMNS, start=1):
        col_letter_ = get_column_letter(col_idx)
        locked = col_name in LEAD_AUTHOR_COLUMNS
        for row in range(1, n_rows + 2):
            ws.cell(row=row, column=col_idx).protection = Protection(locked=locked)

    ws.protection = SheetProtection(
        sheet=True, password=sheet_password,
        formatCells=False, formatColumns=False, formatRows=False,
        insertRows=False, insertColumns=False, deleteRows=False, deleteColumns=False,
        sort=False, autoFilter=False, pivotTables=True,
        selectLockedCells=False, selectUnlockedCells=False,
    )


def load_manifest_rows(manifest_path: str) -> list:
    import csv
    with open(manifest_path, newline="") as f:
        return list(csv.DictReader(f))


def load_packets_by_item_id(evidence_dump_path: str, manifest_rows: list) -> dict:
    """The evidence dump is a JSON array (packets carry only chain/address/anon_id, not the
    real item_id, by blinding design). Reconstruct the item_id mapping by matching on
    (chain, address) -- unique within a manifest -- rather than assuming list order."""
    with open(evidence_dump_path) as f:
        packets = json.load(f)
    by_chain_address = {(p["chain"], p["address"].lower()): p for p in packets}
    result = {}
    for row in manifest_rows:
        key = (row["chain"], row["address"].lower())
        if key not in by_chain_address:
            raise KeyError(f"no evidence packet found for item_id={row['item_id']} ({key})")
        result[row["item_id"]] = by_chain_address[key]
    return result


def load_llm_reviews(reviews_path: str) -> dict:
    with open(reviews_path) as f:
        data = json.load(f)
    return data["reviews"]
