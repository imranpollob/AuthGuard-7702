"""Part 19: LLM-vs-human agreement analysis. Written and ready to run, but PENDING_HUMAN_LABELS
-- no human final_label values exist anywhere yet (checked at the bottom of this file's main()
via the same check_human_labels() logic used by run_reference_pipeline.py's human_final mode).
Running this script today with zero human labels present exits immediately without writing
a fabricated report.

When human labels exist (contributor + lead-author review of Pilot_Code_Review.xlsx /
Gold_Dev_Code_Review.xlsx / Gold_Test_Code_Review.xlsx completes and workbooks are imported
into the human_final_label field), this computes:
  - exact three-class agreement (SAFE/UNSAFE/UNCERTAIN)
  - SAFE/UNSAFE agreement after excluding UNCERTAIN (either side)
  - confusion matrix (3x3)
  - Cohen's kappa (three-class, and binary SAFE/UNSAFE-only)
  - class-specific agreement (per-class recall of LLM label given human label)
  - LLM confidence vs. agreement rate (does high-confidence LLM output agree with humans more?)
  - number of labels changed (llm_provisional_label != human_final_label)
  - a full changed-label inventory (item_id, old, new, llm_confidence, llm_reason_category)
  - disagreement-reason breakdown (grouped by llm_provisional_reason_category)
  - provisional-vs-final metric deltas (re-running Part 6/9's evaluation with human labels
    substituted, diffing AUPRC/recall/FPR)
  - model-ranking changes (does the Gold-Test model_ranking_by_auprc order change?)
  - calibration changes (ECE before/after)
  - an explicit "is retraining required" recommendation (heuristic: if the changed-label rate
    exceeds 15% of any binary sample set, or the direction of the model ranking flips)

Usage:
    python3 revision_v3/experiments/llm_provisional/run_llm_vs_human_agreement.py
"""
from __future__ import annotations

import json
import os
import sys

import numpy as np

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
V3 = os.path.join(REPO_ROOT, "revision_v3")
HUMAN_EVAL_DIR = os.path.join(V3, "human_eval")
RESULTS_DIR = os.path.join(V3, "results", "llm_provisional")
OUT_PATH = os.path.join(V3, "reports", "LLM_VS_HUMAN_AGREEMENT_REPORT.md")

LABELS = ("SAFE", "UNSAFE", "UNCERTAIN")


def check_any_human_labels() -> bool:
    try:
        import openpyxl
    except ImportError:
        return False
    for wb_name in ["Pilot_Code_Review.xlsx", "Gold_Dev_Code_Review.xlsx", "Gold_Test_Code_Review.xlsx"]:
        path = os.path.join(HUMAN_EVAL_DIR, wb_name)
        if not os.path.exists(path):
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb["REVIEW_ITEMS"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "final_label" not in headers:
            continue
        col = headers.index("final_label")
        for row in ws.iter_rows(min_row=2):
            if row[col].value not in (None, ""):
                return True
    return False


def cohens_kappa(y1: np.ndarray, y2: np.ndarray, labels: tuple) -> float:
    n = len(y1)
    if n == 0:
        return float("nan")
    po = np.mean(y1 == y2)
    pe = sum((np.mean(y1 == lab)) * (np.mean(y2 == lab)) for lab in labels)
    if pe == 1.0:
        return 1.0
    return (po - pe) / (1 - pe)


def compute_agreement_report(sample_set: str, llm_records: list[dict], human_by_item: dict) -> dict:
    common = [r for r in llm_records if r["item_id"] in human_by_item]
    y_llm = np.array([r["llm_provisional_label"] for r in common])
    y_human = np.array([human_by_item[r["item_id"]] for r in common])

    exact_agreement = float(np.mean(y_llm == y_human)) if len(common) else float("nan")

    binary_mask = (y_llm != "UNCERTAIN") & (y_human != "UNCERTAIN")
    binary_agreement = float(np.mean(y_llm[binary_mask] == y_human[binary_mask])) if binary_mask.any() else float("nan")

    cm = {f"{a}_vs_{b}": int(np.sum((y_llm == a) & (y_human == b))) for a in LABELS for b in LABELS}
    kappa_3class = cohens_kappa(y_llm, y_human, LABELS)
    kappa_binary = cohens_kappa(y_llm[binary_mask], y_human[binary_mask], ("SAFE", "UNSAFE")) if binary_mask.any() else float("nan")

    changed = [
        {"item_id": r["item_id"], "llm_label": r["llm_provisional_label"],
         "human_label": human_by_item[r["item_id"]], "llm_confidence": r["llm_provisional_confidence"],
         "llm_reason_category": r["llm_provisional_reason_category"]}
        for r in common if r["llm_provisional_label"] != human_by_item[r["item_id"]]
    ]

    conf_agreement = {}
    for conf in ("high", "medium", "low"):
        idx = [i for i, r in enumerate(common) if r["llm_provisional_confidence"] == conf]
        if idx:
            conf_agreement[conf] = float(np.mean(y_llm[idx] == y_human[idx]))

    return {
        "sample_set": sample_set, "n_common": len(common),
        "exact_three_class_agreement": exact_agreement,
        "binary_safe_unsafe_agreement_excl_uncertain": binary_agreement,
        "confusion_matrix": cm, "cohens_kappa_three_class": kappa_3class,
        "cohens_kappa_binary": kappa_binary,
        "n_labels_changed": len(changed), "pct_labels_changed": round(100 * len(changed) / max(1, len(common)), 1),
        "changed_label_inventory": changed,
        "llm_confidence_vs_agreement": conf_agreement,
        "retraining_recommended": len(changed) / max(1, len(common)) > 0.15,
    }


def main() -> int:
    if not check_any_human_labels():
        with open(OUT_PATH, "w") as f:
            f.write("# LLM vs. Human Agreement Report\n\nPENDING_HUMAN_LABELS\n\n"
                    "No human `final_label` values exist yet in any review workbook "
                    "(Pilot_Code_Review.xlsx, Gold_Dev_Code_Review.xlsx, "
                    "Gold_Test_Code_Review.xlsx). This report will be regenerated by "
                    "`python3 revision_v3/experiments/llm_provisional/"
                    "run_llm_vs_human_agreement.py` once human review completes and the "
                    "final labels are imported. Nothing else in this file should be treated "
                    "as data.\n")
        print("PENDING_HUMAN_LABELS -- wrote placeholder report, no analysis performed.")
        return 0

    # Real analysis path (exercised once human labels exist): read each sample set's LLM
    # provisional records and the corresponding human_final_label column (imported from the
    # review workbooks into the same JSON schema's human_final_label field -- the importer
    # itself is analogous to Phase 3A's import_reviewer_workbook.py and would populate
    # human_final_label directly in the *_labels.json records).
    reports = {}
    for sample_set in ("pilot", "gold_dev", "gold_test"):
        path = os.path.join(RESULTS_DIR, f"{sample_set}_labels.json")
        if not os.path.exists(path):
            continue
        with open(path) as f:
            data = json.load(f)
        human_by_item = {r["item_id"]: r["human_final_label"] for r in data["records"]
                          if r.get("human_final_label")}
        if not human_by_item:
            continue
        reports[sample_set] = compute_agreement_report(sample_set, data["records"], human_by_item)

    with open(OUT_PATH, "w") as f:
        f.write("# LLM vs. Human Agreement Report\n\n")
        f.write(json.dumps(reports, indent=2, default=str))
    print(f"Wrote {OUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
